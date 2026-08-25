from __future__ import annotations

import re
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Iterable

from openpyxl import load_workbook

from services.invoice_recognition import (
    InvoiceFile,
    InvoiceRecognitionError,
    normalize_invoice_result,
)

try:
    import xlrd
except Exception:  # pragma: no cover - dependency is optional for old .xls
    xlrd = None


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "article": ("артикул", "код", "номенклатурный номер", "sku"),
    "name": ("наименование", "название", "товар", "номенклатура", "позиция"),
    "quantity": ("количество", "кол-во", "кол во", "qty", "шт"),
    "unit": ("единица", "ед изм", "ед. изм.", "ед"),
    "unit_price_before_discount": (
        "цена до скидки",
        "базовая цена",
        "первоначальная цена",
        "розничная цена",
    ),
    "discount_percent": ("скидка %", "% скидки", "процент скидки"),
    "discount_amount": ("сумма скидки", "скидка руб"),
    "unit_price": (
        "цена со скидкой",
        "цена после скидки",
        "цена продажи",
        "цена за единицу",
        "стоимость ед",
        "цена",
    ),
    "line_total": (
        "сумма со скидкой",
        "сумма после скидки",
        "итог строки",
        "сумма строки",
        "сумма",
        "стоимость",
        "всего",
    ),
}

FINAL_TOTAL_MARKERS = (
    "итого к оплате",
    "всего к оплате",
    "сумма к оплате",
    "к оплате",
    "итого со скидкой",
    "всего со скидкой",
    "сумма с учетом скидки",
    "сумма с учётом скидки",
    "итоговая сумма",
    "итого",
)

INTERMEDIATE_TOTAL_MARKERS = (
    "без скидки",
    "до скидки",
    "сумма скидки",
    "размер скидки",
    "подытог",
    "ндс",
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return " ".join(str(value).replace("\n", " ").split()).strip()


def _normalized(value: Any) -> str:
    text = _text(value).casefold().replace("ё", "е")
    text = re.sub(r"[^a-zа-я0-9%]+", " ", text)
    return " ".join(text.split())


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if result == result and result not in {float("inf"), float("-inf")} else None

    text = _text(value).replace("\xa0", "").replace(" ", "")
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None

    # Поддержка 1 234,56; 1.234,56; 1,234.56 и обычных целых значений.
    comma = text.rfind(",")
    dot = text.rfind(".")
    if comma >= 0 and dot >= 0:
        decimal = "," if comma > dot else "."
        thousands = "." if decimal == "," else ","
        text = text.replace(thousands, "").replace(decimal, ".")
    elif comma >= 0:
        parts = text.split(",")
        if len(parts) == 2 and 1 <= len(parts[1]) <= 4:
            text = parts[0] + "." + parts[1]
        else:
            text = "".join(parts)
    elif dot >= 0 and text.count(".") > 1:
        parts = text.split(".")
        if 1 <= len(parts[-1]) <= 4:
            text = "".join(parts[:-1]) + "." + parts[-1]
        else:
            text = "".join(parts)
    try:
        result = float(text)
        return result if result == result and result not in {float("inf"), float("-inf")} else None
    except (TypeError, ValueError):
        return None


def _load_xlsx(data: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise InvoiceRecognitionError("Не удалось открыть Excel-файл. Проверьте, что он не повреждён.") from exc

    best_rows: list[list[Any]] = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        nonempty = sum(1 for row in rows if any(_text(cell) for cell in row))
        if nonempty > sum(1 for row in best_rows if any(_text(cell) for cell in row)):
            best_rows = rows
    workbook.close()
    return best_rows


def _load_xls(data: bytes) -> list[list[Any]]:
    if xlrd is None:
        raise InvoiceRecognitionError("Для старого формата .xls не установлена библиотека xlrd. Сохраните файл как .xlsx.")
    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        sheet = max(workbook.sheets(), key=lambda item: item.nrows)
        rows = [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows)]
        workbook.release_resources()
        return rows
    except Exception as exc:
        raise InvoiceRecognitionError("Не удалось открыть XLS-файл. Сохраните его как XLSX и отправьте снова.") from exc


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    best_index = -1
    best_columns: dict[str, int] = {}
    best_score = 0
    for row_index, row in enumerate(rows[:40]):
        columns: dict[str, int] = {}
        for column_index, value in enumerate(row):
            header = _normalized(value)
            if not header:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if any(header == alias or header.startswith(alias + " ") for alias in aliases):
                    columns.setdefault(field, column_index)
                    break
        score = len(columns) + (2 if "name" in columns else 0) + (1 if "quantity" in columns else 0)
        if score > best_score:
            best_index, best_columns, best_score = row_index, columns, score

    if best_index < 0 or "name" not in best_columns or not ({"unit_price", "line_total"} & set(best_columns)):
        raise InvoiceRecognitionError(
            "В Excel не найдена таблица накладной. Нужны колонки «Наименование», «Количество» и «Цена» или «Сумма»."
        )
    return best_index, best_columns


def _cell(row: list[Any], columns: dict[str, int], field: str) -> Any:
    index = columns.get(field)
    return row[index] if index is not None and index < len(row) else None


def _find_document_meta(rows: list[list[Any]], header_index: int) -> dict[str, str | None]:
    meta: dict[str, str | None] = {"invoice_number": None, "invoice_date": None, "supplier": None}
    header_rows = rows[:header_index]
    prefix = " ".join(_text(cell) for row in header_rows for cell in row if _text(cell))
    number_match = re.search(r"(?:накладн\w*|сч[её]т\w*)\s*(?:№|N)?\s*([A-Za-zА-Яа-я0-9\-/]+)", prefix, re.I)
    date_match = re.search(r"\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\b", prefix)
    if number_match:
        meta["invoice_number"] = number_match.group(1)
    if date_match:
        meta["invoice_date"] = date_match.group(1)
    for row in header_rows:
        line = " ".join(_text(cell) for cell in row if _text(cell))
        supplier_match = re.search(r"(?:поставщик|продавец)\s*[:\-]\s*(.{3,180})$", line, re.I)
        if supplier_match:
            meta["supplier"] = " ".join(supplier_match.group(1).split())
            break
    return meta


def _find_printed_total(rows: Iterable[list[Any]]) -> float | None:
    candidates: list[tuple[int, int, float]] = []
    for row_index, row in enumerate(rows):
        label = " ".join(_normalized(cell) for cell in row if _text(cell))
        if not label:
            continue
        has_strong_final_marker = any(marker in label for marker in FINAL_TOTAL_MARKERS[:-1])
        if not has_strong_final_marker and any(marker in label for marker in INTERMEDIATE_TOTAL_MARKERS):
            continue
        priority = next((len(FINAL_TOTAL_MARKERS) - index for index, marker in enumerate(FINAL_TOTAL_MARKERS) if marker in label), 0)
        if not priority:
            continue
        numbers = [_number(cell) for cell in row]
        valid = [value for value in numbers if value is not None and value >= 0]
        if valid:
            candidates.append((priority, row_index, valid[-1]))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return round(candidates[0][2], 2)


def recognize_invoice_excel(file: InvoiceFile) -> dict[str, Any]:
    if not file.data:
        raise InvoiceRecognitionError("Excel-файл пустой.")
    if len(file.data) > 20 * 1024 * 1024:
        raise InvoiceRecognitionError("Файл слишком большой. Максимальный размер — 20 МБ.")

    is_xls = file.filename.casefold().endswith(".xls") or file.mime_type == "application/vnd.ms-excel"
    rows = _load_xls(file.data) if is_xls else _load_xlsx(file.data)
    if not rows:
        raise InvoiceRecognitionError("В Excel-файле нет данных.")

    header_index, columns = _find_header(rows)
    meta = _find_document_meta(rows, header_index)
    items: list[dict[str, Any]] = []
    warnings: list[str] = []

    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        name = _text(_cell(row, columns, "name"))
        normalized_name = _normalized(name)
        if not name or any(marker in normalized_name for marker in FINAL_TOTAL_MARKERS):
            continue

        quantity = _number(_cell(row, columns, "quantity"))
        before_price = _number(_cell(row, columns, "unit_price_before_discount"))
        discount_percent = _number(_cell(row, columns, "discount_percent"))
        discount_amount = _number(_cell(row, columns, "discount_amount"))
        unit_price = _number(_cell(row, columns, "unit_price"))
        line_total = _number(_cell(row, columns, "line_total"))
        if quantity is None or quantity <= 0:
            quantity = 1.0
            warnings.append(f"Строка {row_index}: количество не указано, принято значение 1.")

        if unit_price is None and before_price is not None:
            if discount_percent is not None and 0 <= discount_percent < 100:
                unit_price = before_price * (1 - discount_percent / 100)
            elif discount_amount is not None:
                unit_price = (before_price * quantity - discount_amount) / quantity
            else:
                unit_price = before_price

        if unit_price is None and line_total is not None:
            unit_price = line_total / quantity
        if line_total is None and unit_price is not None:
            line_total = quantity * unit_price
        if unit_price is None or line_total is None or unit_price < 0 or line_total < 0:
            continue

        expected = round(quantity * unit_price, 2)
        if abs(line_total - expected) > max(0.02, abs(line_total) * 0.0005):
            if before_price is None and unit_price > line_total / quantity:
                before_price = unit_price
            unit_price = line_total / quantity
            warnings.append(f"Строка {row_index}: цена рассчитана из итоговой суммы строки с учётом возможной скидки.")

        items.append(
            {
                "article": _text(_cell(row, columns, "article")) or None,
                "product_name": name[:500],
                "quantity": round(quantity, 4),
                "unit": _text(_cell(row, columns, "unit")) or "шт",
                "unit_price_before_discount": None if before_price is None else round(before_price, 2),
                "discount_percent": discount_percent,
                "discount_amount": discount_amount,
                "unit_price": round(unit_price, 2),
                "line_total": round(line_total, 2),
            }
        )
        if len(items) >= 150:
            warnings.append("В отчёт взяты первые 150 товарных позиций.")
            break

    if not items:
        raise InvoiceRecognitionError("В Excel не найдено товарных строк с названием и ценой.")

    calculated_total = round(sum(float(item["line_total"]) for item in items), 2)
    printed_total = _find_printed_total(rows[header_index + 1 :])

    raw_result = {
        "document_type": "Накладная / счёт (Excel)",
        "invoice_number": meta["invoice_number"],
        "invoice_date": meta["invoice_date"],
        "supplier": meta["supplier"],
        "buyer": None,
        "responsible_manager": None,
        "currency": "RUB",
        "subtotal_before_discount": calculated_total,
        "discount_amount": None,
        "amount_payable": printed_total,
        "total_amount": printed_total if printed_total is not None else calculated_total,
        "vat_amount": None,
        "confidence": 0.97 if not warnings else 0.9,
        "warnings": list(dict.fromkeys(warnings))[:20],
        "items": items,
    }
    result = normalize_invoice_result(raw_result)
    result["recognition_model"] = "excel"
    result["verification_performed"] = True
    result["recognized_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    return result
