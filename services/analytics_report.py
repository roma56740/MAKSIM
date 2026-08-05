from __future__ import annotations

import os
import tempfile
from typing import Any, List

import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _safe(v: Any) -> Any:
    return "" if v is None else v


def _money(v: Any) -> float:
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _percent(deal_amount: Any, reward_amount: Any) -> float | str:
    deal = _money(deal_amount)
    reward = _money(reward_amount)
    if deal <= 0:
        return ""
    return round(reward / deal * 100, 2)

def _invoice_period_expr(alias: str = "") -> str:
    prefix = f"{alias}." if alias else ""
    return (
        "CASE "
        f"WHEN {prefix}status IN ('approved', 'rejected') "
        f"THEN COALESCE({prefix}handled_at, {prefix}updated_at, {prefix}created_at) "
        f"ELSE {prefix}created_at "
        "END"
    )


def _invoice_period_condition(alias: str = "") -> str:
    expr = _invoice_period_expr(alias)
    return f"{expr} >= ? AND {expr} < ?"


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(55, max(10, max_len + 2))


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    header_font = Font(bold=True)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for r in rows:
        ws.append([_safe(x) for x in r])

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _autosize(ws)


def _tmp_xlsx(prefix: str) -> str:
    fd, path = tempfile.mkstemp(prefix=prefix, suffix=".xlsx")
    os.close(fd)
    return path


async def build_admin_report_xlsx(db_path: str, start_iso: str, end_iso: str, period: str) -> str:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    async with aiosqlite.connect(db_path) as db:
        db.row_factory = aiosqlite.Row

        invoice_where = _invoice_period_condition("")
        invoice_where_i = _invoice_period_condition("i")
        invoice_period_i = _invoice_period_expr("i")

        users_total = (await (await db.execute("SELECT COUNT(*) AS c FROM users")).fetchone())["c"]
        users_new = (await (await db.execute(
            "SELECT COUNT(*) AS c FROM users WHERE created_at >= ? AND created_at < ?",
            (start_iso, end_iso)
        )).fetchone())["c"]

        reg_total = (await (await db.execute(
            "SELECT COUNT(*) AS c FROM registrations WHERE created_at >= ? AND created_at < ?",
            (start_iso, end_iso)
        )).fetchone())["c"]

        inv_total = (await (await db.execute(
            f"SELECT COUNT(*) AS c FROM invoices WHERE {invoice_where}",
            (start_iso, end_iso)
        )).fetchone())["c"]

        po_total = (await (await db.execute(
            "SELECT COUNT(*) AS c FROM payouts WHERE created_at >= ? AND created_at < ?",
            (start_iso, end_iso)
        )).fetchone())["c"]

        inv_sums = await (await db.execute(
            "SELECT COALESCE(SUM(COALESCE(deal_amount,0)),0) AS deal_sum, "
            "COALESCE(SUM(COALESCE(reward_amount,0)),0) AS reward_sum "
            f"FROM invoices WHERE {invoice_where}",
            (start_iso, end_iso)
        )).fetchone()

        po_sums = await (await db.execute(
            "SELECT COALESCE(SUM(COALESCE(amount,0)),0) AS s "
            "FROM payouts WHERE created_at >= ? AND created_at < ?",
            (start_iso, end_iso)
        )).fetchone()

        item_sums = await (await db.execute(
            "SELECT COALESCE(SUM(ii.quantity),0) AS qty, COUNT(ii.id) AS lines "
            "FROM invoice_items ii JOIN invoices i ON i.id = ii.invoice_id "
            f"WHERE i.status='approved' AND {invoice_where_i}",
            (start_iso, end_iso),
        )).fetchone()

        price_uploads = (await (await db.execute(
            "SELECT COUNT(*) AS c FROM supplier_prices WHERE uploaded_at >= ? AND uploaded_at < ?",
            (start_iso, end_iso)
        )).fetchone())["c"]

        ws_summary.append(["Период (ISO)", f"{start_iso} → {end_iso}"])
        ws_summary.append(["Пользователей всего", users_total])
        ws_summary.append(["Новых пользователей за период", users_new])
        ws_summary.append(["Заявок на регистрацию за период", reg_total])
        ws_summary.append(["Накладных за период", inv_total])
        ws_summary.append(["Сумма продаж (₽)", _money(inv_sums["deal_sum"])])
        ws_summary.append(["Сумма вознаграждений (₽)", _money(inv_sums["reward_sum"])])
        ws_summary.append(["Товарных строк в накладных", int(item_sums["lines"] or 0)])
        ws_summary.append(["Продано единиц товара", _money(item_sums["qty"])])
        ws_summary.append(["Запросов на выплаты за период", po_total])
        ws_summary.append(["Сумма выплат (₽)", _money(po_sums["s"])])
        ws_summary.append(["Загружено Excel-прайсов за период", price_uploads])

        for row in ws_summary.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_summary.column_dimensions["A"].width = 34
        ws_summary.column_dimensions["B"].width = 70

        # Пользователи
        ws_users = wb.create_sheet("Пользователи")
        users_rows = await (await db.execute(
            "SELECT tg_id, full_name, phone, reg_type, status, created_at, updated_at FROM users ORDER BY created_at DESC"
        )).fetchall()
        _write_table(
            ws_users,
            ["tg_id", "ФИО", "Телефон", "Тип", "Статус", "Создан", "Обновлен"],
            [[r["tg_id"], r["full_name"], r["phone"], r["reg_type"], r["status"], r["created_at"], r["updated_at"]] for r in users_rows],
        )

        # Заявки (период)
        ws_regs = wb.create_sheet("Заявки")
        regs_rows = await (await db.execute(
            "SELECT tg_id, reg_type, full_name, phone, file_id, file_kind, status, reason, created_at, updated_at "
            "FROM registrations WHERE created_at >= ? AND created_at < ? ORDER BY created_at DESC",
            (start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_regs,
            ["tg_id", "Тип", "ФИО", "Телефон", "file_id", "file_kind", "Статус", "Причина", "Создано", "Обновлено"],
            [[r["tg_id"], r["reg_type"], r["full_name"], r["phone"], r["file_id"], r["file_kind"], r["status"], r["reason"], r["created_at"], r["updated_at"]] for r in regs_rows],
        )

        # Накладные (период)
        ws_inv = wb.create_sheet("Накладные")
        inv_rows = await (await db.execute(
            "SELECT i.id, i.tg_id, u.full_name, i.supplier_id, s.name AS supplier_name, "
            "i.deal_amount, i.reward_amount, i.file_id, i.file_kind, i.comment, i.status, i.reason, "
            "i.created_at, i.handled_at, i.updated_at, "
            f"{invoice_period_i} AS report_date "
            "FROM invoices i "
            "LEFT JOIN users u ON u.tg_id = i.tg_id "
            "LEFT JOIN suppliers s ON s.id = i.supplier_id "
            f"WHERE {invoice_where_i} "
            "ORDER BY report_date DESC, i.id DESC",
            (start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_inv,
            ["id", "tg_id", "Пользователь", "supplier_id", "Поставщик", "Сумма продаж", "Процент", "Вознаграждение", "file_id", "file_kind",
             "Комментарий", "Статус", "Причина", "Создано", "Дата решения", "Обновлено", "Дата отчета"],
            [[r["id"], r["tg_id"], r["full_name"], r["supplier_id"], r["supplier_name"], r["deal_amount"], _percent(r["deal_amount"], r["reward_amount"]), r["reward_amount"],
              r["file_id"], r["file_kind"], r["comment"], r["status"], r["reason"], r["created_at"], r["handled_at"], r["updated_at"], r["report_date"]] for r in inv_rows],
        )


        # Сводная по менеджерам
        ws_mgr = wb.create_sheet("Сводная менеджеры")
        manager_rows = await (await db.execute(
            "WITH inv AS ("
            "  SELECT i.id, i.tg_id, i.deal_amount, i.reward_amount "
            "  FROM invoices i "
            f"  WHERE i.status='approved' AND {invoice_where_i}"
            "), item_agg AS ("
            "  SELECT invoice_id, COUNT(id) AS item_lines, "
            "         COALESCE(SUM(quantity),0) AS items_qty, "
            "         COALESCE(SUM(line_total),0) AS items_sum "
            "  FROM invoice_items GROUP BY invoice_id"
            ") "
            "SELECT inv.tg_id, u.full_name, u.phone, "
            "       COUNT(inv.id) AS invoices_count, "
            "       COALESCE(SUM(inv.deal_amount),0) AS deal_sum, "
            "       COALESCE(SUM(inv.reward_amount),0) AS reward_sum, "
            "       COALESCE(SUM(item_agg.item_lines),0) AS item_lines, "
            "       COALESCE(SUM(item_agg.items_qty),0) AS items_qty, "
            "       COALESCE(SUM(item_agg.items_sum),0) AS items_sum "
            "FROM inv "
            "LEFT JOIN users u ON u.tg_id = inv.tg_id "
            "LEFT JOIN item_agg ON item_agg.invoice_id = inv.id "
            "GROUP BY inv.tg_id, u.full_name, u.phone "
            "ORDER BY deal_sum DESC, reward_sum DESC",
            (start_iso, end_iso),
        )).fetchall()
        _write_table(
            ws_mgr,
            ["tg_id", "Менеджер", "Телефон", "Накладных", "Сумма накладных",
             "Вознаграждение", "Товарных строк", "Продано единиц", "Сумма товаров"],
            [[r["tg_id"], r["full_name"], r["phone"], r["invoices_count"], r["deal_sum"],
              r["reward_sum"], r["item_lines"], r["items_qty"], r["items_sum"]] for r in manager_rows],
        )

        # Сводная одинаковых товаров по каждому менеджеру
        ws_product_summary = wb.create_sheet("Товары менеджеров")
        product_summary_rows = await (await db.execute(
            "SELECT i.tg_id, u.full_name, MIN(ii.product_name) AS product_name, "
            "       COALESCE(SUM(ii.quantity),0) AS quantity, "
            "       MIN(ii.unit_price) AS min_price, MAX(ii.unit_price) AS max_price, "
            "       CASE WHEN SUM(ii.quantity) > 0 "
            "            THEN SUM(ii.line_total) / SUM(ii.quantity) ELSE 0 END AS avg_price, "
            "       COALESCE(SUM(ii.line_total),0) AS total_sum "
            "FROM invoice_items ii "
            "JOIN invoices i ON i.id = ii.invoice_id "
            "LEFT JOIN users u ON u.tg_id = i.tg_id "
            f"WHERE i.status='approved' AND {invoice_where_i} "
            "GROUP BY i.tg_id, COALESCE(NULLIF(ii.product_key, ''), LOWER(TRIM(ii.product_name))) "
            "ORDER BY u.full_name COLLATE NOCASE, total_sum DESC",
            (start_iso, end_iso),
        )).fetchall()
        _write_table(
            ws_product_summary,
            ["tg_id", "Менеджер", "Товар", "Количество", "Мин. цена", "Макс. цена",
             "Средняя цена продажи", "Общая сумма"],
            [[r["tg_id"], r["full_name"], r["product_name"], r["quantity"], r["min_price"],
              r["max_price"], r["avg_price"], r["total_sum"]] for r in product_summary_rows],
        )

        # Подробные товары по каждой накладной
        ws_items = wb.create_sheet("Товары детально")
        item_rows = await (await db.execute(
            "SELECT ii.id, ii.invoice_id, i.tg_id, u.full_name, ii.product_name, "
            "       ii.quantity, ii.unit_price, ii.line_total, i.deal_amount, "
            "       i.reward_amount, i.created_at, i.handled_at, "
            f"       {invoice_period_i} AS report_date "
            "FROM invoice_items ii "
            "JOIN invoices i ON i.id = ii.invoice_id "
            "LEFT JOIN users u ON u.tg_id = i.tg_id "
            f"WHERE i.status='approved' AND {invoice_where_i} "
            "ORDER BY report_date DESC, ii.invoice_id DESC, ii.id ASC",
            (start_iso, end_iso),
        )).fetchall()
        _write_table(
            ws_items,
            ["id", "Накладная", "tg_id", "Менеджер", "Товар", "Количество",
             "Цена продажи", "Сумма товара", "Сумма накладной", "Вознаграждение",
             "Создано", "Дата решения", "Дата отчета"],
            [[r["id"], r["invoice_id"], r["tg_id"], r["full_name"], r["product_name"],
              r["quantity"], r["unit_price"], r["line_total"], r["deal_amount"],
              r["reward_amount"], r["created_at"], r["handled_at"], r["report_date"]] for r in item_rows],
        )

        # Выплаты (период)
        ws_po = wb.create_sheet("Выплаты")
        po_rows = await (await db.execute(
            "SELECT p.id, p.tg_id, u.full_name, p.amount, p.status, p.period_start, p.period_end, "
            "p.comment, p.paid_at, p.created_at, p.updated_at "
            "FROM payouts p LEFT JOIN users u ON u.tg_id = p.tg_id "
            "WHERE p.created_at >= ? AND p.created_at < ? "
            "ORDER BY p.created_at DESC",
            (start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_po,
            ["id", "tg_id", "Пользователь", "Сумма", "Статус", "period_start", "period_end", "Комментарий", "paid_at", "Создано", "Обновлено"],
            [[r["id"], r["tg_id"], r["full_name"], r["amount"], r["status"], r["period_start"], r["period_end"], r["comment"], r["paid_at"], r["created_at"], r["updated_at"]] for r in po_rows],
        )

        # Excel-прайсы (период)
        ws_prices = wb.create_sheet("Excel-прайсы")
        pr_rows = await (await db.execute(
            "SELECT sp.supplier_id, s.name AS supplier_name, sp.file_name, sp.tg_file_id, sp.uploaded_by, sp.uploaded_at "
            "FROM supplier_prices sp LEFT JOIN suppliers s ON s.id = sp.supplier_id "
            "WHERE sp.uploaded_at >= ? AND sp.uploaded_at < ? "
            "ORDER BY sp.uploaded_at DESC",
            (start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_prices,
            ["supplier_id", "Поставщик", "Имя файла", "tg_file_id", "uploaded_by", "uploaded_at"],
            [[r["supplier_id"], r["supplier_name"], r["file_name"], r["tg_file_id"], r["uploaded_by"], r["uploaded_at"]] for r in pr_rows],
        )

        # КП (период)
        ws_kp = wb.create_sheet("КП")
        kp_sess = await (await db.execute(
            "SELECT ks.tg_id, u.full_name, ks.supplier_id, s.name AS supplier_name, ks.created_at, ks.updated_at "
            "FROM kp_sessions ks "
            "LEFT JOIN users u ON u.tg_id = ks.tg_id "
            "LEFT JOIN suppliers s ON s.id = ks.supplier_id "
            "WHERE ks.created_at >= ? AND ks.created_at < ? "
            "ORDER BY ks.created_at DESC",
            (start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_kp,
            ["tg_id", "Пользователь", "supplier_id", "Поставщик", "Создано", "Обновлено"],
            [[r["tg_id"], r["full_name"], r["supplier_id"], r["supplier_name"], r["created_at"], r["updated_at"]] for r in kp_sess],
        )

        ws_kp_items = wb.create_sheet("КП позиции")
        kp_items = await (await db.execute(
            "SELECT k.id, k.tg_id, u.full_name, k.supplier_id, s.name AS supplier_name, "
            "k.product_id, k.title, k.price, k.final_price, k.qty, k.url, k.image_url, k.image_path, k.created_at "
            "FROM kp_items k "
            "LEFT JOIN users u ON u.tg_id = k.tg_id "
            "LEFT JOIN suppliers s ON s.id = k.supplier_id "
            "WHERE k.created_at >= ? AND k.created_at < ? "
            "ORDER BY k.created_at DESC",
            (start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_kp_items,
            ["id", "tg_id", "Пользователь", "supplier_id", "Поставщик", "product_id", "Название",
             "Цена", "Цена фин.", "Кол-во", "url", "image_url", "image_path", "created_at"],
            [[r["id"], r["tg_id"], r["full_name"], r["supplier_id"], r["supplier_name"], r["product_id"], r["title"],
              r["price"], r["final_price"], r["qty"], r["url"], r["image_url"], r["image_path"], r["created_at"]] for r in kp_items],
        )

    path = _tmp_xlsx("admin_report_")
    wb.save(path)
    return path


async def build_user_report_xlsx(db_path: str, tg_id: int, start_iso: str, end_iso: str, period: str) -> str:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    async with aiosqlite.connect(db_path) as db:
        db.row_factory = aiosqlite.Row

        invoice_where_i = _invoice_period_condition("i")
        invoice_period_i = _invoice_period_expr("i")

        user = await (await db.execute(
            "SELECT tg_id, full_name, phone, reg_type, status, created_at FROM users WHERE tg_id = ?",
            (tg_id,)
        )).fetchone()

        inv_sums = await (await db.execute(
            "SELECT COUNT(*) AS cnt, "
            "COALESCE(SUM(COALESCE(i.deal_amount,0)),0) AS deal_sum, "
            "COALESCE(SUM(COALESCE(i.reward_amount,0)),0) AS reward_sum "
            "FROM invoices i "
            f"WHERE i.tg_id = ? AND {invoice_where_i}",
            (tg_id, start_iso, end_iso)
        )).fetchone()

        po_sums = await (await db.execute(
            "SELECT COUNT(*) AS cnt, COALESCE(SUM(COALESCE(amount,0)),0) AS s "
            "FROM payouts WHERE tg_id = ? AND created_at >= ? AND created_at < ?",
            (tg_id, start_iso, end_iso)
        )).fetchone()

        kp_cnt = (await (await db.execute(
            "SELECT COUNT(*) AS c FROM kp_items WHERE tg_id = ? AND created_at >= ? AND created_at < ?",
            (tg_id, start_iso, end_iso)
        )).fetchone())["c"]

        user_item_sums = await (await db.execute(
            "SELECT COALESCE(SUM(ii.quantity),0) AS qty, COUNT(ii.id) AS lines "
            "FROM invoice_items ii JOIN invoices i ON i.id = ii.invoice_id "
            f"WHERE i.tg_id = ? AND i.status='approved' AND {invoice_where_i}",
            (tg_id, start_iso, end_iso),
        )).fetchone()

        ws_summary.append(["Период (ISO)", f"{start_iso} → {end_iso}"])
        ws_summary.append(["tg_id", tg_id])
        ws_summary.append(["ФИО", _safe(user["full_name"]) if user else "—"])
        ws_summary.append(["Телефон", _safe(user["phone"]) if user else "—"])
        ws_summary.append(["Тип", _safe(user["reg_type"]) if user else "—"])
        ws_summary.append(["Статус", _safe(user["status"]) if user else "—"])
        ws_summary.append(["Дата регистрации", _safe(user["created_at"]) if user else "—"])
        ws_summary.append(["Накладных за период", int(inv_sums["cnt"])])
        ws_summary.append(["Сумма продаж (₽)", _money(inv_sums["deal_sum"])])
        ws_summary.append(["Вознаграждение (₽)", _money(inv_sums["reward_sum"])])
        ws_summary.append(["Товарных строк", int(user_item_sums["lines"] or 0)])
        ws_summary.append(["Продано единиц товара", _money(user_item_sums["qty"])])
        ws_summary.append(["Запросов выплат за период", int(po_sums["cnt"])])
        ws_summary.append(["Сумма выплат (₽)", _money(po_sums["s"])])
        ws_summary.append(["Позиции КП за период", int(kp_cnt)])

        for row in ws_summary.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_summary.column_dimensions["A"].width = 34
        ws_summary.column_dimensions["B"].width = 70

        ws_inv = wb.create_sheet("Накладные")
        inv_rows = await (await db.execute(
            "SELECT i.id, i.supplier_id, i.deal_amount, i.reward_amount, i.file_id, i.file_kind, "
            "i.comment, i.status, i.reason, i.created_at, i.handled_at, i.updated_at, "
            f"{invoice_period_i} AS report_date "
            "FROM invoices i "
            f"WHERE i.tg_id = ? AND {invoice_where_i} "
            "ORDER BY report_date DESC, i.id DESC",
            (tg_id, start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_inv,
            ["id", "supplier_id", "Сумма продаж", "Процент", "Вознаграждение", "file_id", "file_kind", "Комментарий", "Статус", "Причина", "Создано", "Дата решения", "Обновлено", "Дата отчета"],
            [[r["id"], r["supplier_id"], r["deal_amount"], _percent(r["deal_amount"], r["reward_amount"]), r["reward_amount"], r["file_id"], r["file_kind"], r["comment"], r["status"], r["reason"], r["created_at"], r["handled_at"], r["updated_at"], r["report_date"]] for r in inv_rows],
        )


        ws_user_items = wb.create_sheet("Товары")
        user_item_rows = await (await db.execute(
            "SELECT ii.invoice_id, ii.product_name, ii.quantity, ii.unit_price, ii.line_total, "
            "       i.deal_amount, i.reward_amount, i.created_at, i.handled_at, "
            f"       {invoice_period_i} AS report_date "
            "FROM invoice_items ii JOIN invoices i ON i.id = ii.invoice_id "
            f"WHERE i.tg_id = ? AND i.status='approved' AND {invoice_where_i} "
            "ORDER BY report_date DESC, ii.invoice_id DESC, ii.id ASC",
            (tg_id, start_iso, end_iso),
        )).fetchall()
        _write_table(
            ws_user_items,
            ["Накладная", "Товар", "Количество", "Цена продажи", "Сумма товара",
             "Сумма накладной", "Вознаграждение", "Создано", "Дата решения", "Дата отчета"],
            [[r["invoice_id"], r["product_name"], r["quantity"], r["unit_price"], r["line_total"],
              r["deal_amount"], r["reward_amount"], r["created_at"], r["handled_at"], r["report_date"]]
             for r in user_item_rows],
        )

        ws_user_product_summary = wb.create_sheet("Сводка товаров")
        user_product_rows = await (await db.execute(
            "SELECT MIN(ii.product_name) AS product_name, SUM(ii.quantity) AS quantity, "
            "       MIN(ii.unit_price) AS min_price, MAX(ii.unit_price) AS max_price, "
            "       CASE WHEN SUM(ii.quantity) > 0 "
            "            THEN SUM(ii.line_total) / SUM(ii.quantity) ELSE 0 END AS avg_price, "
            "       SUM(ii.line_total) AS total_sum "
            "FROM invoice_items ii JOIN invoices i ON i.id = ii.invoice_id "
            f"WHERE i.tg_id = ? AND i.status='approved' AND {invoice_where_i} "
            "GROUP BY COALESCE(NULLIF(ii.product_key, ''), LOWER(TRIM(ii.product_name))) ORDER BY total_sum DESC",
            (tg_id, start_iso, end_iso),
        )).fetchall()
        _write_table(
            ws_user_product_summary,
            ["Товар", "Количество", "Мин. цена", "Макс. цена", "Средняя цена", "Общая сумма"],
            [[r["product_name"], r["quantity"], r["min_price"], r["max_price"],
              r["avg_price"], r["total_sum"]] for r in user_product_rows],
        )

        ws_po = wb.create_sheet("Выплаты")
        po_rows = await (await db.execute(
            "SELECT id, amount, status, period_start, period_end, comment, paid_at, created_at, updated_at "
            "FROM payouts WHERE tg_id = ? AND created_at >= ? AND created_at < ? "
            "ORDER BY created_at DESC",
            (tg_id, start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_po,
            ["id", "Сумма", "Статус", "period_start", "period_end", "Комментарий", "paid_at", "Создано", "Обновлено"],
            [[r["id"], r["amount"], r["status"], r["period_start"], r["period_end"], r["comment"], r["paid_at"], r["created_at"], r["updated_at"]] for r in po_rows],
        )

        ws_kp = wb.create_sheet("КП позиции")
        kp_rows = await (await db.execute(
            "SELECT id, supplier_id, product_id, title, description, price, final_price, url, qty, created_at "
            "FROM kp_items WHERE tg_id = ? AND created_at >= ? AND created_at < ? "
            "ORDER BY created_at DESC",
            (tg_id, start_iso, end_iso)
        )).fetchall()
        _write_table(
            ws_kp,
            ["id", "supplier_id", "product_id", "Название", "Описание", "Цена", "Цена фин.", "url", "Кол-во", "created_at"],
            [[r["id"], r["supplier_id"], r["product_id"], r["title"], r["description"], r["price"], r["final_price"], r["url"], r["qty"], r["created_at"]] for r in kp_rows],
        )

    path = _tmp_xlsx(f"user_{tg_id}_")
    wb.save(path)
    return path
