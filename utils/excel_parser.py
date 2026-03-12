from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

import openpyxl


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v)
    s = s.replace("₽", "").replace("руб.", "").replace(" ", "").replace("\xa0", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = re.sub(r"[^\d\-]", "", str(v))
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


@dataclass
class ExcelRow:
    source_pk: str | None
    code: str
    description: str | None
    price: float | None
    discount_percent: float | None
    final_price: float | None
    product_type: str | None
    stock_qty: int | None
    url: str | None


def parse_products_xlsx(path: str) -> list[ExcelRow]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx: dict[str, int] = {}

    for i, h in enumerate(headers):
        key = _norm(str(h or ""))
        idx[key] = i

    def col(*variants: str) -> int | None:
        for v in variants:
            k = _norm(v)
            if k in idx:
                return idx[k]
        # гибкий поиск
        for k, i in idx.items():
            for v in variants:
                if _norm(v) in k:
                    return i
        return None

    c_pk = col("id товара (pk)", "id товара", "pk")
    c_code = col("код товара", "код", "артикул", "sku")
    c_desc = col("описание")
    c_price = col("цена, ₽", "цена")
    c_disc = col("скидка, %", "процент скидки", "скидка")
    c_final = col("финальная цена, ₽", "финальная цена")
    c_type = col("тип товара", "тип")
    c_stock = col("шт. осталось", "остаток", "кол-во")
    c_url = col("ссылка на товар", "ссылка на товар в магазине", "ссылка")

    if c_code is None:
        raise ValueError("В файле нет колонки 'Код товара'")

    rows: list[ExcelRow] = []

    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c + 1).value for c in range(ws.max_column)]
        code_raw = values[c_code]
        if code_raw is None or str(code_raw).strip() == "":
            continue

        code = str(code_raw).strip()

        source_pk = str(values[c_pk]).strip() if c_pk is not None and values[c_pk] not in (None, "") else None
        desc = str(values[c_desc]).strip() if c_desc is not None and values[c_desc] not in (None, "") else None

        price = _to_float(values[c_price]) if c_price is not None else None
        disc = _to_float(values[c_disc]) if c_disc is not None else None
        final = _to_float(values[c_final]) if c_final is not None else None

        if final is None and price is not None and disc is not None:
            final = round(price * (1 - disc / 100.0), 2)

        ptype = str(values[c_type]).strip() if c_type is not None and values[c_type] not in (None, "") else None
        stock = _to_int(values[c_stock]) if c_stock is not None else None
        url = str(values[c_url]).strip() if c_url is not None and values[c_url] not in (None, "") else None

        rows.append(
            ExcelRow(
                source_pk=source_pk,
                code=code,
                description=desc,
                price=price,
                discount_percent=disc,
                final_price=final,
                product_type=ptype,
                stock_qty=stock,
                url=url,
            )
        )

    return rows
