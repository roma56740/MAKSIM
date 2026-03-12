# handlers/user/catalog.py
from __future__ import annotations

import html
import math
import os
import re
import tempfile
import inspect
from datetime import datetime
from typing import Any, Iterable
import json

from aiogram import F, Router
from aiogram.filters import BaseFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, FSInputFile, InlineKeyboardButton, Message
from aiogram.utils.keyboard import InlineKeyboardBuilder

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from callbacks import UserCatalogCb, UserKpCb
from config import Settings
from db import (
    # users / suppliers / products
    count_products,
    count_suppliers,
    get_product,
    get_supplier,
    get_user,
    is_admin,
    list_products,
    list_suppliers,
    # KP
    add_kp_product,
    add_kp_web_item,
    clear_kp,
    count_kp_items,
    get_kp_session,
    list_kp_items,
    remove_kp_item,
    set_kp_supplier,
)
from keyboards.user import user_back_cancel_kb, user_main_kb
from services.product_enrich import enrich_from_url
from utils.msg_clean import safe_delete

try:
    import aiosqlite
except Exception:  # pragma: no cover
    aiosqlite = None

router = Router()

SUPPLIERS_PAGE_SIZE = 8
PRODUCTS_PAGE_SIZE = 8
SEARCH_PAGE_SIZE = 8
KP_PAGE_SIZE = 6
TOP_K = 1 # кол-во
# результаты поиска по ссылке
SITE_PAGE_SIZE = 8
SITE_MAX_RESULTS = 80

# -------------------- text helpers --------------------

_WORD_RE = re.compile(r"[0-9A-Za-zА-Яа-яЁё]+(?:[-_][0-9A-Za-zА-Яа-яЁё]+)*", re.U)

_STOPWORDS = {
    "и", "в", "во", "на", "по", "для", "с", "со", "к", "ко", "от", "до", "за", "из", "или", "а", "но",
    "the", "and", "or", "for", "with", "to", "of", "in", "on", "by",
    "шт", "штука", "штук", "уп", "упак", "упаковка",
    "ml", "l", "л", "литр", "литра", "литров",
}

# результаты поиска по ссылке
SITE_PAGE_SIZE = 8
SITE_MAX_RESULTS = 80

# новые
SITE_VARIANTS_MAX = 50          # сколько вариантов названия генерим
SITE_MATCHES_STORE_MAX = 4000   # сколько id максимум держим в state (хватает на одного поставщика)


# слова/фразы, которые чаще всего “мусорят” название
_NOISE_PHRASES = [
    "в подарочной упаковке",
    "в подарочной коробке",
    "подарочная упаковка",
    "подарочная коробка",
    "подарочн",
    "упаковк",
    "коробк",
    "gift box",
    "gift",
    "box",
    "set",
    "набор",
]

# категории, которые иногда полезны, но часто лишние — дадим вариант и с ними и без них
_CATEGORY_WORDS = {
    "коньяк", "бренди", "виски", "водка", "джин", "ром", "текила", "ликер", "ликёр", "шампанское",
    "gin", "rum", "tequila", "vodka", "whisky", "whiskey", "cognac", "brandy", "liqueur", "champagne",
}


def _norm_text(s: str | None) -> str:
    t = (s or "").strip()
    t = t.replace("ё", "е").replace("Ё", "Е")
    t = re.sub(r"[^\w\s%]", " ", t, flags=re.UNICODE)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _short(s: str | None, n: int = 40) -> str:
    t = (s or "").strip()
    t = re.sub(r"\s+", " ", t)
    return (t[: n - 1] + "…") if len(t) > n else t


def _money(v: Any) -> str:
    if v is None or v == "":
        return "—"
    try:
        return f"{float(v):,.2f}".replace(",", " ").replace(".", ",")
    except Exception:
        return str(v)


def _fmt_volume(v: Any) -> str:
    if v is None or v == "":
        return "—"
    try:
        x = float(v)
        s = f"{x:g}".replace(".", ",")
        return s
    except Exception:
        return str(v)


def _stock_qty(p: dict[str, Any]) -> int:
    try:
        x = p.get("stock_qty")
        if x is None or x == "":
            return 0
        return int(x)
    except Exception:
        return 0


def _clamp_page(page: int, total: int, page_size: int) -> tuple[int, int]:
    pages = max(1, math.ceil((total or 0) / page_size))
    page = max(0, min(int(page), pages - 1))
    return page, pages


def _nav_buttons(left_cb: str | None, right_cb: str | None) -> list[InlineKeyboardButton]:
    btns: list[InlineKeyboardButton] = []
    if left_cb:
        btns.append(InlineKeyboardButton(text="⬅️", callback_data=left_cb))
    if right_cb:
        btns.append(InlineKeyboardButton(text="➡️", callback_data=right_cb))
    return btns


async def _safe_render_call(call: CallbackQuery, text: str, kb: InlineKeyboardBuilder | None = None) -> None:
    markup = kb.as_markup() if kb else None
    try:
        if call.message:
            await call.message.edit_text(text=text, reply_markup=markup)
    except Exception:
        try:
            if call.message:
                await call.message.delete()
        except Exception:
            pass
        if call.message:
            await call.message.bot.send_message(call.message.chat.id, text, reply_markup=markup)

    try:
        await call.answer()
    except Exception:
        pass


def _chunks(items: list[Any], n: int) -> Iterable[list[Any]]:
    for i in range(0, len(items), n):
        yield items[i : i + n]


# -------------------- Title variants (без ИИ/LLM) --------------------

_CYR_MAP = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ж": "zh", "з": "z", "и": "i", "й": "y",
    "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "",
    "э": "e", "ю": "yu", "я": "ya",
    "ё": "e",
}


def _translit_ru_to_lat(s: str) -> str:
    out = []
    for ch in (s or ""):
        low = ch.lower()
        if low in _CYR_MAP:
            rep = _CYR_MAP[low]
            out.append(rep)
        else:
            out.append(ch)
    t = "".join(out)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _remove_noise_phrases(t: str) -> str:
    s = " " + (t or "").strip() + " "
    low = s.lower()
    for ph in _NOISE_PHRASES:
        if ph in low:
            # грубо вырезаем по lower-индексу
            # (достаточно, чтобы убрать "подарочная упаковка" и похожее)
            low = low.replace(ph, " ")
    # восстановим по low (для упрощения)
    low = re.sub(r"\s+", " ", low).strip()
    return low


def _tokens(t: str) -> list[str]:
    t = _norm_text(t)
    out: list[str] = []
    for w in t.split(" "):
        if not w:
            continue
        w_cf = w.casefold()
        if w_cf in _STOPWORDS:
            continue
        out.append(w_cf)
    return out


def _extract_latin_words(t: str) -> str:
    parts = re.findall(r"[A-Za-z0-9]+(?:[-_][A-Za-z0-9]+)*", t or "")
    return " ".join(parts).strip()


def _make_title_variants(site_title: str) -> list[str]:
    """
    Генерирует варианты запроса из названия страницы:
    - оригинал
    - без мусорных фраз (gift box / подарочная упаковка)
    - без категорий (коньяк/виски/...)
    - только латиница (если есть)
    - короткие варианты (2-4 слова)
    - транслит кириллицы в латиницу
    """
    title = (site_title or "").strip()
    if not title:
        return []

    variants: list[str] = []

    def add(v: str) -> None:
        v = re.sub(r"\s+", " ", (v or "").strip())
        if len(v) < 2:
            return
        if len(v) > 80:
            v = v[:80].strip()
        if v not in variants:
            variants.append(v)

    # 1) оригинал
    add(title)

    # 2) нормализация/очистка
    cleaned = _remove_noise_phrases(title)
    add(cleaned)

    # 3) без категорий
    tok = _tokens(cleaned)
    tok_no_cat = [w for w in tok if w not in _CATEGORY_WORDS]
    if tok_no_cat:
        add(" ".join(tok_no_cat))

    # 4) латиница из оригинала
    latin = _extract_latin_words(title)
    if latin:
        add(latin)

    # 5) латиница из очищенного
    latin2 = _extract_latin_words(cleaned)
    if latin2:
        add(latin2)

    # 6) короткие “ядра”
    tks = [w for w in tok_no_cat if w not in _STOPWORDS]
    if len(tks) >= 2:
        add(" ".join(tks[:2]))
    if len(tks) >= 3:
        add(" ".join(tks[:3]))
    if len(tks) >= 4:
        add(" ".join(tks[:4]))

    # 7) транслит (если много кириллицы)
    tr = _translit_ru_to_lat(title)
    if tr and tr != title:
        add(tr)

    tr2 = _translit_ru_to_lat(" ".join(tok_no_cat))
    if tr2 and tr2 != " ".join(tok_no_cat):
        add(tr2)

    # ограничим количество, чтобы запросы не раздувались
    return variants[:10]

def _dedup_keep_order(items: list[str]) -> list[str]:
    out: list[str] = []
    seen = set()
    for x in items:
        x = re.sub(r"\s+", " ", (x or "").strip())
        if len(x) < 2:
            continue
        key = x.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(x)
    return out


def _expand_title_variants_heuristic(site_title: str, max_variants: int = SITE_VARIANTS_MAX) -> list[str]:
    """
    Быстрый “ИИ-подобный” генератор вариантов без LLM:
    - n-граммы (1..4 слова)
    - версии с/без категории (виски/whisky/…)
    - версии без объёма/цифр
    - латиница, транслит
    - короткие ядра (бренд/модель/ключевое слово)
    Итого обычно 30–80 → режем до max_variants.
    """
    title = (site_title or "").strip()
    if not title:
        return []

    base10 = _make_title_variants(title)  # твоя текущая функция (до 10)

    cleaned = _remove_noise_phrases(title)
    tok = _tokens(cleaned)
    tok_no_cat = [w for w in tok if w not in _CATEGORY_WORDS]

    # убираем “объёмные” токены типа 07, 0_7, 700, 750, 1000, ml, l, л...
    def is_volume_token(t: str) -> bool:
        t = (t or "").casefold()
        if t in {"ml", "l", "л"}:
            return True
        if re.fullmatch(r"\d{2,4}", t):  # 500 700 750 1000 ...
            return True
        if re.fullmatch(r"\d+[,.]\d+", t):  # 0.7 0,7
            return True
        return False

    tok_no_vol = [t for t in tok_no_cat if not is_volume_token(t)]
    tok_no_nums = [t for t in tok_no_cat if not re.search(r"\d", t)]

    variants: list[str] = []
    variants.extend(base10)
    variants.append(title)
    variants.append(cleaned)
    if tok_no_cat:
        variants.append(" ".join(tok_no_cat))
    if tok_no_vol:
        variants.append(" ".join(tok_no_vol))
    if tok_no_nums:
        variants.append(" ".join(tok_no_nums))

    # латиница / транслит
    latin = _extract_latin_words(title)
    if latin:
        variants.append(latin)
    latin2 = _extract_latin_words(cleaned)
    if latin2:
        variants.append(latin2)

    tr = _translit_ru_to_lat(title)
    if tr and tr != title:
        variants.append(tr)

    tr2 = _translit_ru_to_lat(" ".join(tok_no_cat))
    if tr2 and tr2 != " ".join(tok_no_cat):
        variants.append(tr2)

    # n-граммы (ядра) — из tok_no_cat
    def add_ngrams(tokens: list[str], max_n: int = 4) -> None:
        if not tokens:
            return
        for n in range(1, max_n + 1):
            for i in range(0, len(tokens) - n + 1):
                variants.append(" ".join(tokens[i:i+n]))

    add_ngrams(tok_no_cat, 4)
    add_ngrams(tok_no_vol, 4)

    # самые “сильные” одиночные токены (бренд/ключевые) — длиннее 5
    for t in tok_no_cat:
        if len(t) >= 6:
            variants.append(t)

    out = _dedup_keep_order(variants)
    return out[:max_variants]


async def _make_title_variants_llm(settings: Settings, site_title: str, max_variants: int = SITE_VARIANTS_MAX) -> list[str]:
    """
    Опционально: если есть OpenAI ключ — попросим модель выдать JSON-массив до 50 вариантов.
    Если ключа/модуля нет — вернём [] и всё продолжит работать на эвристиках.
    """
    title = (site_title or "").strip()
    if not title:
        return []

    api_key = getattr(settings, "openai_api_key", None) or os.getenv("OPENAI_API_KEY")
    if not api_key:
        return []

    try:
        from openai import AsyncOpenAI
    except Exception:
        return []

    model = (
        getattr(settings, "openai_model", None)
        or getattr(settings, "gpt_model", None)
        or os.getenv("OPENAI_MODEL")
        or "gpt-4o-mini"
    )

    prompt = (
        "Сгенерируй варианты поисковых запросов для товара по названию со страницы.\n"
        f"Название: {title}\n\n"
        f"Нужно: JSON-массив строк (только JSON, без текста), до {max_variants} вариантов.\n"
        "Правила:\n"
        "- Варианты 1–6 слов, без мусора.\n"
        "- Дай варианты на русском и английском (если уместно).\n"
        "- Дай варианты без категории (например без слов 'виски/whisky', 'водка/vodka' и т.п.).\n"
        "- Дай короткие ядра (бренд, модель, ключевое слово), включая 1 слово.\n"
        "- Не дублируй одинаковые варианты.\n"
    )

    try:
        client = AsyncOpenAI(api_key=api_key)
        resp = await client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "Ты помощник для генерации вариантов поисковых запросов."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
        )
        text = (resp.choices[0].message.content or "").strip()
        # вытащим JSON-список
        if "[" in text and "]" in text:
            text = text[text.find("["): text.rfind("]") + 1]
        data = json.loads(text)
        if not isinstance(data, list):
            return []
        variants = [str(x) for x in data if isinstance(x, (str, int, float))]
        variants = [str(v).strip() for v in variants if str(v).strip()]
        return _dedup_keep_order(variants)[:max_variants]
    except Exception:
        return []


async def _make_title_variants_smart(settings: Settings, site_title: str, max_variants: int = SITE_VARIANTS_MAX) -> list[str]:
    """
    Итоговый генератор: LLM (если доступен) + эвристики.
    На выходе до ~50 вариантов.
    """
    heur = _expand_title_variants_heuristic(site_title, max_variants=max_variants)
    llm = await _make_title_variants_llm(settings, site_title, max_variants=max_variants)

    merged = _dedup_keep_order([site_title] + llm + heur)
    return merged[:max_variants]

# -------------------- FTS / SQL Search (обычный поисковик) --------------------

_fts_ready = False


def _fts_supported() -> bool:
    return aiosqlite is not None


def _fts_query_from_text(q: str) -> str:
    """
    Превращаем текст в FTS5 MATCH:
    token* AND token*
    """
    tks = _tokens(q)
    # ограничим, чтобы MATCH не был огромным
    tks = tks[:6]
    safe: list[str] = []
    for w in tks:
        # FTS5 не любит спецсимволы — оставим только буквы/цифры/подчёркивания
        ww = re.sub(r"[^\w]", "", w)
        if not ww:
            continue
        safe.append(f"{ww}*")
    return " AND ".join(safe) if safe else ""


async def _ensure_fts(settings: Settings) -> bool:
    """
    Создаёт FTS5 (если возможно) и триггеры.
    Работает как “ускоритель” для обычного поиска.
    """
    global _fts_ready
    if _fts_ready:
        return True
    if not _fts_supported():
        return False

    try:
        async with aiosqlite.connect(settings.db_path) as db:
            await db.execute(
                """
                CREATE VIRTUAL TABLE IF NOT EXISTS product_fts
                USING fts5(
                    title,
                    description,
                    code,
                    content='products',
                    content_rowid='id'
                );
                """
            )

            # триггеры синхронизации
            await db.execute(
                """
                CREATE TRIGGER IF NOT EXISTS products_ai AFTER INSERT ON products BEGIN
                  INSERT INTO product_fts(rowid, title, description, code)
                  VALUES (new.id, COALESCE(new.title,''), COALESCE(new.description,''), COALESCE(new.code,''));
                END;
                """
            )
            await db.execute(
                """
                CREATE TRIGGER IF NOT EXISTS products_ad AFTER DELETE ON products BEGIN
                  INSERT INTO product_fts(product_fts, rowid, title, description, code)
                  VALUES('delete', old.id, old.title, old.description, old.code);
                END;
                """
            )
            await db.execute(
                """
                CREATE TRIGGER IF NOT EXISTS products_au AFTER UPDATE ON products BEGIN
                  INSERT INTO product_fts(product_fts, rowid, title, description, code)
                  VALUES('delete', old.id, old.title, old.description, old.code);
                  INSERT INTO product_fts(rowid, title, description, code)
                  VALUES (new.id, COALESCE(new.title,''), COALESCE(new.description,''), COALESCE(new.code,''));
                END;
                """
            )

            # первичная индексация существующих строк
            # (rebuild бывает недоступен в некоторых сборках — поэтому в try)
            try:
                await db.execute("INSERT INTO product_fts(product_fts) VALUES('rebuild');")
            except Exception:
                pass

            await db.commit()
        _fts_ready = True
        return True
    except Exception:
        return False


async def _search_single_query(
    *,
    settings: Settings,
    supplier_id: int,
    query: str,
    limit: int,
    offset: int,
    in_stock_only: bool = False,
) -> tuple[int, list[dict[str, Any]]]:
    """
    Обычный поиск (как поисковик):
    - сначала FTS5 (если есть)
    - иначе LIKE
    """
    q = (query or "").strip()
    if not q:
        return 0, []

    # ---------- FTS ----------
    if await _ensure_fts(settings):
        match = _fts_query_from_text(q)
        if match:
            stock_clause = "AND COALESCE(p.stock_qty,0) > 0" if in_stock_only else ""
            async with aiosqlite.connect(settings.db_path) as db:
                db.row_factory = aiosqlite.Row

                cur = await db.execute(
                    f"""
                    SELECT COUNT(1)
                    FROM product_fts
                    JOIN products p ON p.id = product_fts.rowid
                    WHERE p.supplier_id = ?
                      AND product_fts MATCH ?
                      {stock_clause}
                    """,
                    (int(supplier_id), match),
                )
                row = await cur.fetchone()
                await cur.close()
                total = int(row[0] or 0) if row else 0

                if total <= 0:
                    return 0, []

                cur = await db.execute(
                    f"""
                    SELECT p.*
                    FROM product_fts
                    JOIN products p ON p.id = product_fts.rowid
                    WHERE p.supplier_id = ?
                      AND product_fts MATCH ?
                      {stock_clause}
                    ORDER BY bm25(product_fts) ASC, COALESCE(p.stock_qty,0) DESC, p.id DESC
                    LIMIT ? OFFSET ?
                    """,
                    (int(supplier_id), match, int(limit), int(offset)),
                )
                rows = await cur.fetchall()
                await cur.close()
                return total, [dict(r) for r in rows]


    # ---------- LIKE fallback ----------
    if aiosqlite is None:
        # совсем простой fallback через имеющиеся функции
        total = await count_products(settings.db_path, supplier_id)
        if total <= 0:
            return 0, []
        # достанем по кускам и отфильтруем (на 4000 норм)
        out: list[dict[str, Any]] = []
        q_cf = _norm_text(q).casefold()
        step = 500
        for off in range(0, total, step):
            batch = await list_products(settings.db_path, supplier_id=supplier_id, limit=step, offset=off)
            for p in batch:
                hay = f"{p.get('title','')} {p.get('description','')} {p.get('code','')}"
                if q_cf in _norm_text(hay).casefold():
                    if (not in_stock_only) or (_stock_qty(p) > 0):
                        out.append(p)
        total2 = len(out)
        return total2, out[offset : offset + limit]

    # SQL LIKE (быстрее, чем питон-фильтр)
    tks = _tokens(q)[:6]
    if not tks:
        return 0, []

    stock_clause = "AND COALESCE(stock_qty,0) > 0" if in_stock_only else ""

    # каждый токен должен встречаться в (title/description/code)
    where_parts: list[str] = []
    params: list[Any] = [int(supplier_id)]
    for w in tks:
        like = f"%{w}%"
        where_parts.append("(LOWER(COALESCE(title,'')) LIKE ? OR LOWER(COALESCE(description,'')) LIKE ? OR LOWER(COALESCE(code,'')) LIKE ?)")
        params += [like, like, like]

    where_sql = " AND ".join(where_parts)

    async with aiosqlite.connect(settings.db_path) as db:
        db.row_factory = aiosqlite.Row

        cur = await db.execute(
            f"""
            SELECT COUNT(1)
            FROM products
            WHERE supplier_id = ?
              {stock_clause}
              AND {where_sql}
            """,
            params,
        )
        row = await cur.fetchone()
        await cur.close()
        total = int(row[0] or 0) if row else 0
        if total <= 0:
            return 0, []

        cur = await db.execute(
            f"""
            SELECT *
            FROM products
            WHERE supplier_id = ?
              {stock_clause}
              AND {where_sql}
            ORDER BY COALESCE(stock_qty,0) DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            params + [int(limit), int(offset)],
        )
        rows = await cur.fetchall()
        await cur.close()
        return total, [dict(r) for r in rows]

async def _rank_products_by_variants(
    *,
    settings: Settings,
    supplier_id: int,
    variants: list[str],
    in_stock_only: bool,
) -> list[int]:
    """
    Очень быстрый ранжированный поиск по поставщику:
    - грузим товары поставщика одним проходом
    - проверяем вхождения токенов/фраз (LIKE-подобно: '%token%')
    - считаем score и сортируем
    Возвращаем список product_id в порядке релевантности.
    """
    variants = _dedup_keep_order([v for v in (variants or []) if v and v.strip()])[:SITE_VARIANTS_MAX]
    if not variants:
        return []

    phrases = [_norm_text(v).casefold() for v in variants]
    vtoks = [_tokens(v)[:5] for v in variants]

    rows: list[dict[str, Any]] = []

    if aiosqlite is not None:
        async with aiosqlite.connect(settings.db_path) as db:
            db.row_factory = aiosqlite.Row
            cur = await db.execute(
                """
                SELECT id, title, description, code, strength, product_type, volume, stock_qty
                FROM products
                WHERE supplier_id = ?
                """,
                (int(supplier_id),),
            )
            fetched = await cur.fetchall()
            await cur.close()
            rows = [dict(r) for r in fetched]
    else:
        total = await count_products(settings.db_path, supplier_id)
        step = 500
        for off in range(0, total, step):
            batch = await list_products(settings.db_path, supplier_id=supplier_id, limit=step, offset=off)
            rows.extend(batch)

    scored: list[tuple[int, int, int]] = []  # (score, stock, id)

    for p in rows:
        stock = _stock_qty(p)
        if in_stock_only and stock <= 0:
            continue

        hay = _norm_text(f"{p.get('title','')} {p.get('description','')} {p.get('code','')}")
        hay_cf = hay.casefold()
        if not hay_cf:
            continue

        best = 0
        # 50 вариантов максимум — быстро
        for ph, tks in zip(phrases, vtoks):
            if not ph and not tks:
                continue

            if ph and ph in hay_cf:
                score = 500 + min(120, len(ph))
            else:
                if not tks:
                    continue
                hits = sum(1 for tk in tks if tk and tk in hay_cf)
                if hits == 0:
                    continue
                if hits == len(tks):
                    score = 350 + hits * 10
                elif hits >= max(1, len(tks) - 1):
                    score = 250 + hits * 10
                else:
                    score = 140 + hits * 10

            if score > best:
                best = score

        if best > 0:
            scored.append((best, stock, int(p["id"])))

    scored.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
    ids = [pid for _, _, pid in scored]
    return ids[:SITE_MATCHES_STORE_MAX]


async def _fetch_products_brief_by_ids(settings: Settings, ids: list[int]) -> list[dict[str, Any]]:
    """
    Достаём минимальные поля для списка-кнопок (как в поиске), сохраняя порядок ids.
    """
    ids = [int(x) for x in ids if int(x) > 0]
    if not ids:
        return []

    if aiosqlite is None:
        out: list[dict[str, Any]] = []
        for pid in ids:
            p = await get_product(settings.db_path, pid)
            if p:
                out.append(p)
        return out

    ph = ",".join(["?"] * len(ids))
    # order by case
    case = "CASE id " + " ".join([f"WHEN ? THEN {i}" for i, _ in enumerate(ids)]) + " END"

    async with aiosqlite.connect(settings.db_path) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute(
            f"""
            SELECT id, title, description, code, strength, product_type, volume, stock_qty
            FROM products
            WHERE id IN ({ph})
            ORDER BY {case}
            """,
            ids + ids,
        )
        rows = await cur.fetchall()
        await cur.close()
        return [dict(r) for r in rows]


async def _search_by_variants(
    *,
    settings: Settings,
    supplier_id: int,
    variants: list[str],
    limit: int,
    offset: int,
    in_stock_only: bool = True,
) -> tuple[int, list[dict[str, Any]]]:
    """
    Поиск “как поисковик”, но запрос = OR(варианты).
    Максимально быстро: 1 SQL запрос.
    """
    variants = [v.strip() for v in (variants or []) if v and v.strip()]
    if not variants:
        return 0, []

    # ---------- FTS ----------
    if await _ensure_fts(settings):
        parts: list[str] = []
        for v in variants[:6]:
            mq = _fts_query_from_text(v)
            if mq:
                parts.append(f"({mq})")
        match = " OR ".join(parts)
        if match:
            stock_clause = "AND COALESCE(p.stock_qty,0) > 0" if in_stock_only else ""
            async with aiosqlite.connect(settings.db_path) as db:
                db.row_factory = aiosqlite.Row

                cur = await db.execute(
                    f"""
                    SELECT COUNT(1)
                    FROM product_fts
                    JOIN products p ON p.id = product_fts.rowid
                    WHERE p.supplier_id = ?
                      AND product_fts MATCH ?
                      {stock_clause}
                    """,
                    (int(supplier_id), match),
                )
                row = await cur.fetchone()
                await cur.close()
                total = int(row[0] or 0) if row else 0
                if total <= 0:
                    return 0, []

                cur = await db.execute(
                    f"""
                    SELECT p.*
                    FROM product_fts
                    JOIN products p ON p.id = product_fts.rowid
                    WHERE p.supplier_id = ?
                      AND product_fts MATCH ?
                      {stock_clause}
                    ORDER BY bm25(product_fts) ASC, COALESCE(p.stock_qty,0) DESC, p.id DESC
                    LIMIT ? OFFSET ?
                    """,
                    (int(supplier_id), match, int(limit), int(offset)),
                )
                rows = await cur.fetchall()
                await cur.close()
                return total, [dict(r) for r in rows]


    # ---------- LIKE fallback ----------
    # ограничим до 6 вариантов (иначе OR может быть очень большой)
    vs = variants[:6]

    if aiosqlite is None:
        # fallback на питон
        total_all = await count_products(settings.db_path, supplier_id)
        if total_all <= 0:
            return 0, []
        out: list[dict[str, Any]] = []
        vs_cf = [_norm_text(v).casefold() for v in vs]
        step = 500
        for off in range(0, total_all, step):
            batch = await list_products(settings.db_path, supplier_id=supplier_id, limit=step, offset=off)
            for p in batch:
                if in_stock_only and _stock_qty(p) <= 0:
                    continue
                hay = f"{p.get('title','')} {p.get('description','')} {p.get('code','')}"
                hay_cf = _norm_text(hay).casefold()
                if any(v in hay_cf for v in vs_cf):
                    out.append(p)
        total2 = len(out)
        return total2, out[offset : offset + limit]

    stock_clause = "AND COALESCE(stock_qty,0) > 0" if in_stock_only else ""

    # OR по вариантам, но каждый вариант = AND по его токенам (чтобы не было “слишком широко”)
    or_parts: list[str] = []
    params: list[Any] = [int(supplier_id)]
    for v in vs:
        tks = _tokens(v)[:5]
        if not tks:
            continue
        and_parts: list[str] = []
        for w in tks:
            like = f"%{w}%"
            and_parts.append("(LOWER(COALESCE(title,'')) LIKE ? OR LOWER(COALESCE(description,'')) LIKE ? OR LOWER(COALESCE(code,'')) LIKE ?)")
            params += [like, like, like]
        or_parts.append("(" + " AND ".join(and_parts) + ")")

    if not or_parts:
        return 0, []

    where_sql = " OR ".join(or_parts)

    async with aiosqlite.connect(settings.db_path) as db:
        db.row_factory = aiosqlite.Row

        cur = await db.execute(
            f"""
            SELECT COUNT(1)
            FROM products
            WHERE supplier_id = ?
              {stock_clause}
              AND ({where_sql})
            """,
            params,
        )
        row = await cur.fetchone()
        await cur.close()
        total = int(row[0] or 0) if row else 0
        if total <= 0:
            return 0, []

        cur = await db.execute(
            f"""
            SELECT *
            FROM products
            WHERE supplier_id = ?
              {stock_clause}
              AND ({where_sql})
            ORDER BY COALESCE(stock_qty,0) DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            params + [int(limit), int(offset)],
        )
        rows = await cur.fetchall()
        await cur.close()
        return total, [dict(r) for r in rows]


# -------------------- Photos (без дефолтной) --------------------

async def _send_product_photo(message: Message, p: dict[str, Any]) -> None:
    try:
        img_path = p.get("image_path")
        if img_path and os.path.exists(img_path):
            await message.answer_photo(FSInputFile(img_path))
            return
    except Exception:
        pass

    try:
        img_url = p.get("image_url")
        if img_url:
            await message.answer_photo(str(img_url))
            return
    except Exception:
        pass


async def _send_site_photo_only(message: Message, site: dict[str, Any]) -> None:
    img_url = (site.get("image_url") or "").strip() if isinstance(site.get("image_url"), str) else ""
    if not img_url:
        return
    try:
        await message.answer_photo(img_url, caption="📷 Фото товара (со страницы)")
    except Exception:
        await message.answer("⚠️ Не получилось отправить фото товара, но продолжу поиск в базе.")


# -------------------- Product UI --------------------

def _product_title_line(p: dict[str, Any]) -> str:
    title = (p.get("title") or p.get("description") or "").strip()
    code = (p.get("code") or "").strip()
    strength = (p.get("strength") or p.get("product_type") or "").strip()

    volume = p.get("volume")
    vol_s = _fmt_volume(volume)

    parts = []
    if title:
        parts.append(title)
    elif code:
        parts.append(code)
    else:
        parts.append("Товар")

    if volume is not None and volume != "":
        parts.append(f"{vol_s} л")
    if strength:
        parts.append(strength)

    line = re.sub(r"\s+", " ", " ".join(parts)).strip()[:120]
    stock = _stock_qty(p)
    badge = "✅" if stock > 0 else "❌"
    return f"{badge} {line}"


def _product_caption(p: dict[str, Any]) -> str:
    title = p.get("title") or None
    strength = p.get("strength") or None
    volume = p.get("volume")

    desc = p.get("description") or None
    product_type = p.get("product_type") or None

    code = p.get("code") or "—"
    source_pk = p.get("source_pk") or "—"

    price = _money(p.get("price"))
    disc = p.get("discount_percent")
    disc_s = f"{_money(disc)}%" if disc is not None else "—"
    final_price = _money(p.get("final_price"))

    stock = _stock_qty(p)
    stock_line = f"✅ <b>{stock} шт</b>" if stock > 0 else "❌ <b>Нет в наличии</b>"

    lines = [
        "🧾 <b>Карточка товара</b>",
        "",
        f"Код товара: <code>{html.escape(str(code))}</code>",
        f"Артикул: <code>{html.escape(str(source_pk))}</code>",
        "",
        f"Номенклатура: <b>{html.escape(_short(title or desc or '—', 180))}</b>",
        f"Крепость/тип: <b>{html.escape(_short(strength or product_type or '—', 80))}</b>",
        f"Объём: <b>{html.escape(_fmt_volume(volume))}</b>",
        "",
        f"Наличие: {stock_line}",
        "",
        f"Цена: <b>{html.escape(price)}</b>",
        f"Скидка: <b>{html.escape(disc_s)}</b>",
        f"Цена для клиента: <b>{html.escape(final_price)}</b>",
    ]

    url = p.get("url")
    if url:
        lines.append("")
        lines.append(f"Ссылка: {html.escape(str(url))}")

    text = "\n".join(lines)
    if len(text) > 3900:
        text = text[:3890] + "…"
    return text


# -------------------- Excel export --------------------

def _safe_filename(name: str, max_len: int = 40) -> str:
    name = re.sub(r"[^\w\s\-\(\)\.]", "", name, flags=re.UNICODE).strip()
    name = re.sub(r"\s+", "_", name)
    if not name:
        name = "supplier"
    return name[:max_len]


def _build_supplier_catalog_xlsx(*, supplier: dict[str, Any], products: list[dict[str, Any]]) -> str:
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "Поставщик"

    ws_info.append(["Поле", "Значение"])
    ws_info.append(["Название", supplier.get("name") or "—"])
    ws_info.append(["Сайт", supplier.get("website") or "—"])
    ws_info.append(["Почта", supplier.get("email") or "—"])
    ws_info.append(["Телефон", supplier.get("phone") or "—"])
    ws_info.append(["Описание", supplier.get("description") or "—"])
    ws_info.append(["Товаров", len(products)])

    ws_info.freeze_panes = "A2"
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 70

    ws = wb.create_sheet("Каталог")
    headers = [
        "ID",
        "Код",
        "Номенклатура",
        "Крепость/Тип",
        "Объём (л)",
        "Наличие (шт)",
        "Цена",
        "Скидка (%)",
        "Цена для клиента",
        "URL",
        "Описание",
        "Image URL",
    ]
    ws.append(headers)

    for p in products:
        ws.append([
            p.get("id"),
            p.get("code"),
            p.get("title") or "",
            p.get("strength") or (p.get("product_type") or ""),
            p.get("volume"),
            p.get("stock_qty"),
            p.get("price"),
            p.get("discount_percent"),
            p.get("final_price"),
            p.get("url") or "",
            p.get("description") or "",
            p.get("image_url") or "",
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = [8, 16, 50, 18, 12, 14, 12, 12, 16, 45, 60, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    stamp = datetime.now().strftime("%Y-%m-%d")
    fname = f"catalog_{_safe_filename(supplier.get('name') or 'supplier')}_{stamp}.xlsx"

    tmp_dir = tempfile.mkdtemp(prefix="supp_catalog_")
    out_path = os.path.join(tmp_dir, fname)
    wb.save(out_path)
    return out_path


# -------------------- Filters / FSM --------------------

class IsApprovedUser(BaseFilter):
    async def __call__(self, event: Message | CallbackQuery, settings: Settings) -> bool:
        uid = event.from_user.id
        if await is_admin(settings.db_path, uid, settings.admin_ids):
            return False
        user = await get_user(settings.db_path, uid)
        return bool(user and user.get("status") == "approved")


class SiteAddForm(StatesGroup):
    waiting_url = State()
    browsing = State()


class ProductSearchForm(StatesGroup):
    waiting_query = State()
    browsing = State()


# -------------------- KP helpers --------------------

async def _supplier_name(settings: Settings, supplier_id: int | None) -> str:
    if not supplier_id:
        return "—"
    supp = await get_supplier(settings.db_path, int(supplier_id))
    return supp["name"] if supp else f"#{supplier_id}"


async def _ensure_single_supplier(settings: Settings, tg_id: int, supplier_id: int) -> tuple[bool, int | None]:
    """
    ✅ Теперь одно КП может содержать позиции разных поставщиков.
    Оставлено для совместимости: больше НЕ ограничиваем поставщика.
    """
    return True, None



async def _kp_has_product(settings: Settings, tg_id: int, product_id: int) -> bool:
    total = await count_kp_items(settings.db_path, tg_id)
    if total <= 0:
        return False
    items = await list_kp_items(settings.db_path, tg_id, limit=total, offset=0)
    return any(int(it.get("product_id") or 0) == int(product_id) for it in items)


async def _add_kp_product_once(settings: Settings, tg_id: int, supplier_id: int, product_id: int) -> tuple[bool, str]:
    p = await get_product(settings.db_path, product_id)
    if not p:
        return False, "⚠️ Товар не найден в базе."

    if await _kp_has_product(settings, tg_id, product_id):
        return False, "⚠️ Этот товар уже добавлен в КП."

    await add_kp_product(settings.db_path, tg_id, supplier_id=supplier_id, product_id=product_id)

    stock = _stock_qty(p)
    if stock <= 0:
        return True, "✅ Добавлено в КП, но товара нет в наличии (0 шт)."

    return True, "✅ Товар добавлен в КП."


async def _save_web_item_best_effort(settings: Settings, tg_id: int, supplier_id: int, site: dict[str, Any]) -> None:
    """
    Сохраняем “как на сайте” (best-effort под разные сигнатуры add_kp_web_item).
    """
    try:
        sig = inspect.signature(add_kp_web_item)
        kwargs: dict[str, Any] = {}
        for name in sig.parameters.keys():
            if name in {"db_path"}:
                kwargs[name] = settings.db_path
            elif name in {"tg_id", "user_id"}:
                kwargs[name] = tg_id
            elif name in {"supplier_id"}:
                kwargs[name] = supplier_id
            elif name in {"title", "name"}:
                kwargs[name] = site.get("title") or ""
            elif name in {"url", "link"}:
                kwargs[name] = site.get("url") or ""
            elif name in {"description"}:
                kwargs[name] = site.get("description") or ""
            elif name in {"product_type"}:
                kwargs[name] = site.get("product_type") or ""
            elif name in {"strength"}:
                kwargs[name] = site.get("strength") or ""
            elif name in {"volume"}:
                kwargs[name] = site.get("volume")
            elif name in {"image_url", "image"}:
                kwargs[name] = site.get("image_url") or ""
        # если у функции позиционные — попробуем самый частый порядок
        if kwargs:
            await add_kp_web_item(**kwargs)  # type: ignore
        else:
            # fallback: db_path, tg_id, supplier_id, title, url, image_url
            await add_kp_web_item(
                settings.db_path, tg_id, supplier_id,
                site.get("title") or "",
                site.get("url") or "",
                site.get("image_url") or "",
            )  # type: ignore
    except Exception:
        # не ломаем flow, даже если сохранение недоступно
        return


# -------------------- Screens: Suppliers --------------------

async def _screen_suppliers(target: Message | CallbackQuery, settings: Settings, page: int = 0) -> None:
    total = await count_suppliers(settings.db_path)
    page, pages = _clamp_page(page, total, SUPPLIERS_PAGE_SIZE)

    items = await list_suppliers(settings.db_path, limit=SUPPLIERS_PAGE_SIZE, offset=page * SUPPLIERS_PAGE_SIZE)

    if total == 0:
        text = "🏢 <b>Поставщики</b>\n\nПока нет поставщиков."
        kb = InlineKeyboardBuilder()
        kb.row(InlineKeyboardButton(text="📄 Моё КП", callback_data=UserKpCb(action="view", page=0, item_id=0).pack()))
        if isinstance(target, CallbackQuery):
            await _safe_render_call(target, text, kb)
        else:
            await target.answer(text, reply_markup=kb.as_markup())
        return

    text = "🏢 <b>Поставщики</b>\nВыберите поставщика:"
    if pages > 1:
        text += f"\n\nСтр. <b>{page + 1}</b> / <b>{pages}</b>"

    kb = InlineKeyboardBuilder()
    for s in items:
        sid = int(s["id"])
        name = s["name"]
 
        kb.add(
            InlineKeyboardButton(
                text=name,
                callback_data=UserCatalogCb(action="supp_open", page=page, supplier_id=sid, product_id=0).pack(),
            )
        )
    kb.adjust(1)

    left_cb = UserCatalogCb(action="supp_page", page=page - 1, supplier_id=0, product_id=0).pack() if page > 0 else None
    right_cb = UserCatalogCb(action="supp_page", page=page + 1, supplier_id=0, product_id=0).pack() if page < pages - 1 else None
    nav = _nav_buttons(left_cb, right_cb)
    if nav:
        kb.row(*nav)

    kb.row(InlineKeyboardButton(text="📄 Моё КП", callback_data=UserKpCb(action="view", page=0, item_id=0).pack()))

    if isinstance(target, CallbackQuery):
        await _safe_render_call(target, text, kb)
    else:
        await target.answer(text, reply_markup=kb.as_markup())


async def _screen_supplier_menu(call: CallbackQuery, settings: Settings, supplier_id: int) -> None:
    supp = await get_supplier(settings.db_path, supplier_id)
    if not supp:
        await call.answer("Поставщик не найден", show_alert=True)
        return

    total = await count_products(settings.db_path, supplier_id)

    name = html.escape(str(supp.get("name") or "—"))
    website = html.escape(str(supp.get("website") or "—"))
    email = html.escape(str(supp.get("email") or "—"))
    phone = html.escape(str(supp.get("phone") or "—"))
    desc = (supp.get("description") or "").strip()
    desc = re.sub(r"\s+", " ", desc)
    desc_short = html.escape(_short(desc, 280)) if desc else "—"

    text = (
        f"🏢 <b>{name}</b>\n"
        f"🌐 Сайт: {website}\n"
        f"✉️ Почта: {email}\n"
        f"📞 Телефон: {phone}\n"
        f"📝 Описание: {desc_short}\n\n"
        f"📦 Товаров в боте: <b>{total}</b>\n\n"
        "Выберите действие:"
    )

    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text="📦 Товары", callback_data=UserCatalogCb(action="prod_page", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="📊 Каталог в Excel", callback_data=UserCatalogCb(action="excel", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="🔎 Поиск по базе", callback_data=UserCatalogCb(action="search_start", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="🔗 Проверить товар по ссылке", callback_data=UserCatalogCb(action="site_add", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="📄 Моё КП", callback_data=UserKpCb(action="view", page=0, item_id=0).pack()))
    kb.row(InlineKeyboardButton(text="⬅️ К поставщикам", callback_data=UserCatalogCb(action="supp_page", page=0, supplier_id=0, product_id=0).pack()))
    await _safe_render_call(call, text, kb)


# -------------------- Excel handler --------------------

@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "excel"))
async def supplier_catalog_excel(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()

    supplier_id = int(callback_data.supplier_id)
    supp = await get_supplier(settings.db_path, supplier_id)
    if not supp:
        await call.answer("Поставщик не найден", show_alert=True)
        return

    total = await count_products(settings.db_path, supplier_id)
    if total <= 0:
        await call.answer("У поставщика нет товаров", show_alert=True)
        await _screen_supplier_menu(call, settings, supplier_id=supplier_id)
        return

    try:
        await call.answer("Готовлю Excel…")
    except Exception:
        pass

    products: list[dict[str, Any]] = []
    limit = 1000
    offset = 0
    while offset < total:
        batch = await list_products(settings.db_path, supplier_id=supplier_id, limit=limit, offset=offset)
        if not batch:
            break
        products.extend(batch)
        offset += limit

    path = None
    try:
        path = _build_supplier_catalog_xlsx(supplier=supp, products=products)
        file = FSInputFile(path)
        caption = (
            f"📊 <b>Каталог в Excel</b>\n"
            f"🏢 Поставщик: <b>{html.escape(str(supp.get('name') or '—'))}</b>\n"
            f"📦 Товаров: <b>{len(products)}</b>"
        )
        if call.message:
            await call.message.answer_document(file, caption=caption)
    finally:
        if path:
            try:
                dir_ = os.path.dirname(path)
                os.remove(path)
                try:
                    os.rmdir(dir_)
                except Exception:
                    pass
            except Exception:
                pass

    await _screen_supplier_menu(call, settings, supplier_id=supplier_id)


# -------------------- Screens: Products --------------------

async def _screen_products(call: CallbackQuery, settings: Settings, supplier_id: int, page: int) -> None:
    total = await count_products(settings.db_path, supplier_id)
    page, pages = _clamp_page(page, total, PRODUCTS_PAGE_SIZE)

    supp = await get_supplier(settings.db_path, supplier_id)
    supp_name = supp["name"] if supp else f"#{supplier_id}"

    if total == 0:
        text = f"📦 <b>{html.escape(str(supp_name))}</b>\n\nПока нет товаров."
        kb = InlineKeyboardBuilder()
        kb.row(InlineKeyboardButton(text="📄 Моё КП", callback_data=UserKpCb(action="view", page=0, item_id=0).pack()))
        kb.row(InlineKeyboardButton(text="⬅️ Назад", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack()))
        await _safe_render_call(call, text, kb)
        return

    items = await list_products(settings.db_path, supplier_id=supplier_id, limit=PRODUCTS_PAGE_SIZE, offset=page * PRODUCTS_PAGE_SIZE)

    text = f"📦 <b>{html.escape(str(supp_name))}</b>\nВыберите товар:"
    if pages > 1:
        text += f"\n\nСтр. <b>{page + 1}</b> / <b>{pages}</b>"

    kb = InlineKeyboardBuilder()
    for p in items:
        pid = int(p["id"])
        title = _short(_product_title_line(p), 56)
        kb.add(
            InlineKeyboardButton(
                text=title,
                callback_data=UserCatalogCb(action="prod_open", page=page, supplier_id=supplier_id, product_id=pid).pack(),
            )
        )
    kb.adjust(1)

    left_cb = UserCatalogCb(action="prod_page", page=page - 1, supplier_id=supplier_id, product_id=0).pack() if page > 0 else None
    right_cb = UserCatalogCb(action="prod_page", page=page + 1, supplier_id=supplier_id, product_id=0).pack() if page < pages - 1 else None
    nav = _nav_buttons(left_cb, right_cb)
    if nav:
        kb.row(*nav)

    kb.row(InlineKeyboardButton(text="🔎 Поиск по базе", callback_data=UserCatalogCb(action="search_start", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="🔗 Проверить товар по ссылке", callback_data=UserCatalogCb(action="site_add", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="📄 Моё КП", callback_data=UserKpCb(action="view", page=0, item_id=0).pack()))
    kb.row(InlineKeyboardButton(text="⬅️ Назад", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack()))
    await _safe_render_call(call, text, kb)


async def _screen_product_card(
    call: CallbackQuery,
    settings: Settings,
    supplier_id: int,
    product_id: int,
    back_page: int,
    back_mode: str = "list",  # "list" | "search" | "site"
) -> None:
    p = await get_product(settings.db_path, product_id)
    if not p:
        await call.answer("Товар не найден", show_alert=True)
        return

    if call.message:
        await _send_product_photo(call.message, p)

    stock = _stock_qty(p)

    kb = InlineKeyboardBuilder()

    btn_text = "➕ Добавить в КП" if stock > 0 else "➕ Добавить в КП (нет в наличии)"
    kb.row(
        InlineKeyboardButton(
            text=btn_text,
            callback_data=UserCatalogCb(action="add", page=back_page, supplier_id=supplier_id, product_id=product_id).pack(),
        )
    )


    kb.row(InlineKeyboardButton(text="📄 Моё КП", callback_data=UserKpCb(action="view", page=0, item_id=0).pack()))

    if back_mode == "search":
        kb.row(
            InlineKeyboardButton(
                text="⬅️ К результатам поиска",
                callback_data=UserCatalogCb(action="search_page", page=back_page, supplier_id=supplier_id, product_id=0).pack(),
            )
        )
    elif back_mode == "site":
        kb.row(
            InlineKeyboardButton(
                text="⬅️ К результатам проверки",
                callback_data=UserCatalogCb(action="site_page", page=back_page, supplier_id=supplier_id, product_id=0).pack(),
            )
        )
    else:
        kb.row(
            InlineKeyboardButton(
                text="⬅️ К списку",
                callback_data=UserCatalogCb(action="prod_page", page=back_page, supplier_id=supplier_id, product_id=0).pack(),
            )
        )

    kb.row(
        InlineKeyboardButton(
            text="🏢 Меню поставщика",
            callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack(),
        )
    )

    await _safe_render_call(call, _product_caption(p), kb)


# -------------------- Entry points (Reply keyboard) --------------------

@router.message(IsApprovedUser(), F.text == "📦 Каталог")
async def user_catalog(message: Message, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_suppliers(message, settings, page=0)


@router.message(IsApprovedUser(), F.text == "📄 Моё КП")
async def user_my_kp(message: Message, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_kp(message, settings, page=0)


# -------------------- Suppliers navigation (Inline) --------------------

@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "supp_page"))
async def supp_page(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_suppliers(call, settings, page=int(callback_data.page))


@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "supp_open"))
async def supp_open(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_supplier_menu(call, settings, supplier_id=int(callback_data.supplier_id))


# -------------------- Products navigation (Inline) --------------------

@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "prod_page"))
async def prod_page(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_products(call, settings, supplier_id=int(callback_data.supplier_id), page=int(callback_data.page))


@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "prod_open"))
async def prod_open(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_product_card(
        call,
        settings,
        supplier_id=int(callback_data.supplier_id),
        product_id=int(callback_data.product_id),
        back_page=int(callback_data.page),
        back_mode="list",
    )


# -------------------- Search flow (обычный поисковик, без ИИ) --------------------

@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "search_start"))
async def search_start(call: CallbackQuery, callback_data: UserCatalogCb, state: FSMContext, settings: Settings) -> None:
    await state.clear()
    supplier_id = int(callback_data.supplier_id)

    await state.update_data(search_supplier_id=supplier_id)
    await state.set_state(ProductSearchForm.waiting_query)

    try:
        if call.message:
            await call.message.delete()
    except Exception:
        pass

    prompt = await call.message.bot.send_message(
        call.message.chat.id,
        "🔎 <b>Поиск по базе</b>\n\n"
        "Напишите запрос (как в обычном поиске):\n"
        "пример: <code>Hennessy Paradis</code> или <code>0,7</code>.\n\n"
        "❌ Отмена — кнопкой ниже.",
        reply_markup=user_back_cancel_kb(),
    )
    await state.update_data(search_prompt_id=prompt.message_id)
    await call.answer()


@router.message(IsApprovedUser(), ProductSearchForm.waiting_query, F.text.in_(["❌ Отмена", "⬅️ Назад"]))
async def search_cancel(message: Message, state: FSMContext, settings: Settings) -> None:
    data = await state.get_data()
    pid = int(data.get("search_prompt_id") or 0)
    if pid:
        await safe_delete(message.bot, message.chat.id, pid)
    await safe_delete(message.bot, message.chat.id, message.message_id)

    supplier_id = int(data.get("search_supplier_id") or 0)
    await state.clear()

    if supplier_id:
        kb = InlineKeyboardBuilder()
        kb.button(
            text="🏢 Меню поставщика",
            callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack(),
        )
        kb.adjust(1)
        await message.answer("✅ Ок, поиск отменён.", reply_markup=kb.as_markup())
    else:
        await message.answer("✅ Ок.", reply_markup=user_main_kb())


@router.message(IsApprovedUser(), ProductSearchForm.waiting_query, F.text)
async def search_got_query(message: Message, state: FSMContext, settings: Settings) -> None:
    q = (message.text or "").strip()
    if len(q) < 2:
        await message.answer("⚠️ Запрос слишком короткий. Напишите минимум 2 символа.", reply_markup=user_back_cancel_kb())
        return

    data = await state.get_data()
    supplier_id = int(data.get("search_supplier_id") or 0)
    prompt_id = int(data.get("search_prompt_id") or 0)

    if prompt_id:
        await safe_delete(message.bot, message.chat.id, prompt_id)
    await safe_delete(message.bot, message.chat.id, message.message_id)

    if not supplier_id:
        await state.clear()
        await message.answer("⚠️ Ошибка: не выбран поставщик. Откройте каталог заново.", reply_markup=user_main_kb())
        return

    await state.update_data(search_query=q)
    await state.set_state(ProductSearchForm.browsing)

    total, items = await _search_single_query(
        settings=settings,
        supplier_id=supplier_id,
        query=q,
        limit=SEARCH_PAGE_SIZE,
        offset=0,
        in_stock_only=False,
    )

    if total == 0:
        kb = InlineKeyboardBuilder()
        kb.button(text="🔁 Новый поиск", callback_data=UserCatalogCb(action="search_start", page=0, supplier_id=supplier_id, product_id=0).pack())
        kb.button(text="⬅️ Назад", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack())
        kb.adjust(1)
        await message.answer(
            "❌ Ничего не нашёл в базе.\n\nПопробуйте другой запрос.",
            reply_markup=kb.as_markup(),
        )
        return

    pages = max(1, math.ceil(total / SEARCH_PAGE_SIZE))
    supp = await get_supplier(settings.db_path, supplier_id)
    supp_name = supp["name"] if supp else f"#{supplier_id}"

    text = (
        f"🔎 <b>Поиск по базе</b> — <b>{html.escape(str(supp_name))}</b>\n"
        f"Запрос: <code>{html.escape(_short(q, 60))}</code>\n"
        f"Найдено: <b>{total}</b>\n"
    )
    if pages > 1:
        text += f"\nСтр. <b>1</b> / <b>{pages}</b>"

    kb = InlineKeyboardBuilder()
    for p in items:
        pid = int(p["id"])
        title = _short(_product_title_line(p), 56)
        kb.add(
            InlineKeyboardButton(
                text=title,
                callback_data=UserCatalogCb(action="search_open", page=0, supplier_id=supplier_id, product_id=pid).pack(),
            )
        )
    kb.adjust(1)

    if pages > 1:
        kb.row(*_nav_buttons(None, UserCatalogCb(action="search_page", page=1, supplier_id=supplier_id, product_id=0).pack()))

    kb.row(InlineKeyboardButton(text="🔁 Новый поиск", callback_data=UserCatalogCb(action="search_start", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="⬅️ Назад", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack()))
    await message.answer(text, reply_markup=kb.as_markup())


@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "search_page"))
async def search_page(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    data = await state.get_data()
    q = (data.get("search_query") or "").strip()
    supplier_id = int(callback_data.supplier_id)
    if not q:
        await call.answer("Поиск не активен. Нажмите «🔎 Поиск» заново.", show_alert=True)
        return

    await state.set_state(ProductSearchForm.browsing)

    total, _ = await _search_single_query(
        settings=settings,
        supplier_id=supplier_id,
        query=q,
        limit=1,
        offset=0,
        in_stock_only=False,
    )
    page, pages = _clamp_page(int(callback_data.page), total, SEARCH_PAGE_SIZE)

    _, items = await _search_single_query(
        settings=settings,
        supplier_id=supplier_id,
        query=q,
        limit=SEARCH_PAGE_SIZE,
        offset=page * SEARCH_PAGE_SIZE,
        in_stock_only=False,
    )

    supp = await get_supplier(settings.db_path, supplier_id)
    supp_name = supp["name"] if supp else f"#{supplier_id}"

    text = (
        f"🔎 <b>Поиск по базе</b> — <b>{html.escape(str(supp_name))}</b>\n"
        f"Запрос: <code>{html.escape(_short(q, 60))}</code>\n"
        f"Найдено: <b>{total}</b>\n"
    )
    if pages > 1:
        text += f"\nСтр. <b>{page + 1}</b> / <b>{pages}</b>"

    kb = InlineKeyboardBuilder()
    for p in items:
        pid = int(p["id"])
        title = _short(_product_title_line(p), 56)
        kb.add(InlineKeyboardButton(text=title, callback_data=UserCatalogCb(action="search_open", page=page, supplier_id=supplier_id, product_id=pid).pack()))
    kb.adjust(1)

    left_cb = UserCatalogCb(action="search_page", page=page - 1, supplier_id=supplier_id, product_id=0).pack() if page > 0 else None
    right_cb = UserCatalogCb(action="search_page", page=page + 1, supplier_id=supplier_id, product_id=0).pack() if page < pages - 1 else None
    nav = _nav_buttons(left_cb, right_cb)
    if nav:
        kb.row(*nav)

    kb.row(InlineKeyboardButton(text="🔁 Новый поиск", callback_data=UserCatalogCb(action="search_start", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="🏢 Меню поставщика", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack()))
    await _safe_render_call(call, text, kb)


@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "search_open"))
async def search_open(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    await state.set_state(ProductSearchForm.browsing)
    await _screen_product_card(
        call,
        settings,
        supplier_id=int(callback_data.supplier_id),
        product_id=int(callback_data.product_id),
        back_page=int(callback_data.page),
        back_mode="search",
    )


# -------------------- Add to KP --------------------

@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "add"))
async def add_to_kp(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings) -> None:
    import os

    tg_id = call.from_user.id
    supplier_id = int(callback_data.supplier_id)

    # DEBUG: чтобы понять, какая БД реально используется
    try:
        await call.answer(f"Успешно добавлено!", show_alert=True)
    except Exception:
        pass

    added, msg = await _add_kp_product_once(
        settings,
        tg_id=tg_id,
        supplier_id=supplier_id,
        product_id=int(callback_data.product_id),
    )
    await call.answer(msg, show_alert=not added)


# -------------------- “Проверка по ссылке” (варианты названия -> поиск как в поисковике) --------------------

async def _safe_enrich(url: str) -> dict[str, Any]:
    """
    Берём данные со страницы (без LLM). Если enrich упал — возвращаем минимум.
    """
    try:
        info = await enrich_from_url(url)
        if isinstance(info, dict):
            return info
    except Exception:
        pass
    return {
        "title": "",
        "description": "",
        "product_type": "",
        "strength": "",
        "volume": None,
        "image_url": "",
        "image": "",
        "url": url,
    }


@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "site_add"))
async def site_add_start(call: CallbackQuery, callback_data: UserCatalogCb, state: FSMContext, settings: Settings) -> None:
    await state.clear()
    tg_id = call.from_user.id
    supplier_id = int(callback_data.supplier_id)

 

    await state.update_data(supplier_id=supplier_id)
    await state.set_state(SiteAddForm.waiting_url)

    try:
        if call.message:
            await call.message.delete()
    except Exception:
        pass

    prompt = await call.message.bot.send_message(
        call.message.chat.id,
        "🔗 <b>Проверка товара по ссылке</b>\n\n"
        "Пришлите ссылку на товар.\n"
        "Я возьму <b>название</b> и сохраню данные со страницы, "
        "затем найду совпадения в базе через обычный поиск (несколько вариантов названия).\n\n"
        "❌ Отмена — кнопкой ниже.",
        reply_markup=user_back_cancel_kb(),
    )
    await state.update_data(site_prompt_id=prompt.message_id)
    await call.answer()


@router.message(IsApprovedUser(), SiteAddForm.waiting_url, F.text.in_(["⬅️ Назад", "❌ Отмена"]))
async def site_add_cancel(message: Message, state: FSMContext, settings: Settings) -> None:
    data = await state.get_data()
    pid = int(data.get("site_prompt_id") or 0)
    if pid:
        await safe_delete(message.bot, message.chat.id, pid)
    await safe_delete(message.bot, message.chat.id, message.message_id)

    supplier_id = int(data.get("supplier_id") or 0)
    await state.clear()

    if supplier_id:
        kb = InlineKeyboardBuilder()
        kb.button(text="🏢 Меню поставщика", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack())
        kb.adjust(1)
        await message.answer("✅ Ок, отменено.", reply_markup=kb.as_markup())
    else:
        await message.answer("✅ Ок.", reply_markup=user_main_kb())


@router.message(IsApprovedUser(), SiteAddForm.waiting_url, F.text)
async def site_add_got_url(message: Message, state: FSMContext, settings: Settings) -> None:
    raw = (message.text or "").strip()

    data = await state.get_data()
    supplier_id = int(data.get("supplier_id") or 0)
    prompt_id = int(data.get("site_prompt_id") or 0)

    if prompt_id:
        await safe_delete(message.bot, message.chat.id, prompt_id)
    await safe_delete(message.bot, message.chat.id, message.message_id)

    if not supplier_id:
        await state.clear()
        await message.answer("⚠️ Ошибка: не выбран поставщик. Откройте каталог заново.", reply_markup=user_main_kb())
        return

    if not re.match(r"^https?://", raw, flags=re.I):
        await message.answer("⚠️ Пришлите именно ссылку (http/https).", reply_markup=user_back_cancel_kb())
        await state.set_state(SiteAddForm.waiting_url)
        return

    url = raw.strip()

    progress = await message.bot.send_message(message.chat.id, "⏳ Читаю страницу…")
    progress_id = progress.message_id

    info = await _safe_enrich(url)
    await safe_delete(message.bot, message.chat.id, progress_id)

    site = {
        "url": url,
        "title": (info.get("title") or "").strip(),
        "description": (info.get("description") or "").strip(),
        "product_type": (info.get("product_type") or "").strip(),
        "strength": (info.get("strength") or "").strip(),
        "volume": info.get("volume"),
        "image_url": (info.get("image_url") or info.get("image") or "").strip()
        if isinstance(info.get("image_url") or info.get("image"), str)
        else "",
    }

    # 1) Сохраняем “как на сайте”
    await _save_web_item_best_effort(settings, message.from_user.id, supplier_id, site)

    # 2) Фото со страницы (если есть)
    if site.get("image_url"):
        await _send_site_photo_only(message, site)

    # 3) Название со страницы (как ты хочешь)
    title = (site.get("title") or site.get("description") or "").strip()
    if title:
        await message.answer(f"🧾 Название со страницы: <b>{html.escape(_short(title, 220))}</b>")
    else:
        await message.answer("🧾 Название со страницы: <b>—</b>")

    # 4) Генерим ~50 вариантов (эвристики + LLM для вариантов, если доступно)
    variants = await _make_title_variants_smart(settings, title, max_variants=SITE_VARIANTS_MAX)
    if not variants and title:
        variants = [title]

    # 5) Быстрый поиск по БД (получаем кандидатов)
    progress2 = await message.bot.send_message(message.chat.id, "⏳ Ищу совпадения в базе…")
    progress2_id = progress2.message_id

    ids = await _rank_products_by_variants(
        settings=settings,
        supplier_id=supplier_id,
        variants=variants,
        in_stock_only=True,
    )

    in_stock_used = True
    if not ids:
        in_stock_used = False
        ids = await _rank_products_by_variants(
            settings=settings,
            supplier_id=supplier_id,
            variants=variants,
            in_stock_only=False,
        )

    await safe_delete(message.bot, message.chat.id, progress2_id)

    if not ids:
        kb = InlineKeyboardBuilder()
        kb.row(InlineKeyboardButton(text="🔗 Другая ссылка", callback_data=UserCatalogCb(action="site_add", page=0, supplier_id=supplier_id, product_id=0).pack()))
        kb.row(InlineKeyboardButton(text="🔎 Поиск по базе", callback_data=UserCatalogCb(action="search_start", page=0, supplier_id=supplier_id, product_id=0).pack()))
        kb.row(InlineKeyboardButton(text="🏢 Меню поставщика", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack()))
        kb.adjust(1)
        await message.answer(
            "🔗 <b>Проверка по ссылке</b>\n\n"
            "❌ <b>В базе не нашёл подходящих товаров</b>\n"
            "Попробуйте другую ссылку или воспользуйтесь поиском по базе.",
            reply_markup=kb.as_markup(),
        )
        await state.clear()
        return

    # -------------------- ВОТ ТУТ: ИИ выбирает лучшие совпадения --------------------

    TOP_K = 5  # поставь 1, если хочешь всегда 1 результат

    # ограничим кандидатов, чтобы промпт не раздувался
    candidate_ids = ids[: min(len(ids), SITE_MAX_RESULTS)]
    candidates = await _fetch_products_brief_by_ids(settings, candidate_ids)

    async def _pick_best_ids_with_llm(
        *,
        site_title: str,
        site_volume: Any,
        site_strength: str,
        site_type: str,
        candidates_brief: list[dict[str, Any]],
        top_k: int,
    ) -> list[int]:
        api_key = getattr(settings, "openai_api_key", None) or os.getenv("OPENAI_API_KEY")
        if not api_key:
            return []

        try:
            from openai import AsyncOpenAI
        except Exception:
            return []

        model = (
            getattr(settings, "openai_model", None)
            or getattr(settings, "gpt_model", None)
            or os.getenv("OPENAI_MODEL")
            or "gpt-4o-mini"
        )

        # максимально компактно, но информативно
        payload = []
        for p in candidates_brief:
            pid = int(p.get("id") or 0)
            if not pid:
                continue
            strength = (p.get("strength") or p.get("product_type") or "") or ""
            payload.append({
                "id": pid,
                "title": _short(str(p.get("title") or p.get("description") or ""), 120),
                "code": str(p.get("code") or ""),
                "volume": p.get("volume"),
                "strength": _short(str(strength), 60),
                "stock_qty": p.get("stock_qty"),
            })

        if not payload:
            return []

        prompt = (
            "Ты выбираешь наиболее похожие товары из базы для товара со страницы.\n"
            "Верни ТОЛЬКО JSON-массив чисел (id товаров), без пояснений.\n\n"
            f"Товар со страницы:\n"
            f"- title: {site_title}\n"
            f"- volume: {site_volume}\n"
            f"- strength: {site_strength}\n"
            f"- type: {site_type}\n\n"
            f"Список кандидатов (выбирай только из них):\n"
            f"{json.dumps(payload, ensure_ascii=False)}\n\n"
            f"Правила:\n"
            f"- верни до {top_k} id, отсортированных от лучшего к худшему;\n"
            f"- если нет нормальных совпадений — верни [];\n"
            f"- не придумывай id.\n"
        )

        try:
            client = AsyncOpenAI(api_key=api_key)
            resp = await client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "Ты помогаешь выбрать самые похожие товары из списка кандидатов."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.0,
            )
            text = (resp.choices[0].message.content or "").strip()
            # вытащим JSON-массив
            if "[" in text and "]" in text:
                text = text[text.find("["): text.rfind("]") + 1]
            arr = json.loads(text)
            if not isinstance(arr, list):
                return []

            allowed = {int(x["id"]) for x in payload if isinstance(x.get("id"), int)}
            out: list[int] = []
            seen = set()
            for x in arr:
                try:
                    pid = int(x)
                except Exception:
                    continue
                if pid in allowed and pid not in seen:
                    seen.add(pid)
                    out.append(pid)
                if len(out) >= top_k:
                    break
            return out
        except Exception:
            return []

    progress3 = await message.bot.send_message(message.chat.id, "🧠 Уточняю совпадение с помощью ИИ…")
    progress3_id = progress3.message_id

    picked_ids = await _pick_best_ids_with_llm(
        site_title=title,
        site_volume=site.get("volume"),
        site_strength=site.get("strength") or "",
        site_type=site.get("product_type") or "",
        candidates_brief=candidates,
        top_k=TOP_K,
    )

    await safe_delete(message.bot, message.chat.id, progress3_id)

    # если ИИ ничего не выбрал — покажем первые TOP_K по нашему скорингу
    final_ids = picked_ids if picked_ids else candidate_ids[:TOP_K]
    final_ids = [int(x) for x in final_ids if int(x) > 0]

    total = len(final_ids)
    items = await _fetch_products_brief_by_ids(settings, final_ids[:SITE_PAGE_SIZE])

    # сохраним результаты, чтобы открыть карточку/листать (обычно страница одна)
    await state.set_state(SiteAddForm.browsing)
    await state.update_data(
        site_supplier_id=supplier_id,
        site_in_stock_used=in_stock_used,
        site_title=title,
        site_variants_preview=(variants[:10] if variants else []),
        site_match_ids=final_ids,      # показываем только то, что выбрал ИИ
        site_all_match_ids=ids,        # на будущее (если захочешь кнопку "показать все")
    )

    supp = await get_supplier(settings.db_path, supplier_id)
    supp_name = supp["name"] if supp else f"#{supplier_id}"

    header = (
        f"🔗 <b>Лучшие совпадения (ИИ)</b> — <b>{html.escape(str(supp_name))}</b>\n"
        f"Название: <code>{html.escape(_short(title or '—', 80))}</code>\n"
        f"Показано: <b>{total}</b>"
    )
    if not in_stock_used:
        header += "\n(включая товары без наличия)"

    kb = InlineKeyboardBuilder()
    for p in items:
        pid = int(p["id"])
        title_btn = _short(_product_title_line(p), 56)
        kb.add(
            InlineKeyboardButton(
                text=title_btn,
                callback_data=UserCatalogCb(action="site_open", page=0, supplier_id=supplier_id, product_id=pid).pack(),
            )
        )
    kb.adjust(1)

    kb.row(InlineKeyboardButton(text="🔗 Другая ссылка", callback_data=UserCatalogCb(action="site_add", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="🏢 Меню поставщика", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack()))

    await message.answer(header, reply_markup=kb.as_markup())


@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "site_page"))
async def site_page(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    data = await state.get_data()
    supplier_id = int(callback_data.supplier_id)

    ids: list[int] = list(data.get("site_match_ids") or [])
    in_stock_used: bool = bool(data.get("site_in_stock_used", True))
    title: str = str(data.get("site_title") or "").strip()

    if not ids:
        await call.answer("Проверка по ссылке устарела. Запустите заново.", show_alert=True)
        await _screen_supplier_menu(call, settings, supplier_id=supplier_id)
        return

    total = len(ids)
    page, pages = _clamp_page(int(callback_data.page), total, SITE_PAGE_SIZE)

    page_ids = ids[page * SITE_PAGE_SIZE : page * SITE_PAGE_SIZE + SITE_PAGE_SIZE]
    items = await _fetch_products_brief_by_ids(settings, page_ids)

    supp = await get_supplier(settings.db_path, supplier_id)
    supp_name = supp["name"] if supp else f"#{supplier_id}"

    text = (
        f"🔗 <b>Совпадения в базе</b> — <b>{html.escape(str(supp_name))}</b>\n"
        f"Название: <code>{html.escape(_short(title or '—', 80))}</code>\n"
        f"Найдено: <b>{total}</b>"
    )
    if not in_stock_used:
        text += "\n(включая товары без наличия)"
    if pages > 1:
        text += f"\n\nСтр. <b>{page + 1}</b> / <b>{pages}</b>"

    kb = InlineKeyboardBuilder()
    for p in items:
        pid = int(p["id"])
        title_btn = _short(_product_title_line(p), 56)
        kb.add(
            InlineKeyboardButton(
                text=title_btn,
                callback_data=UserCatalogCb(action="site_open", page=page, supplier_id=supplier_id, product_id=pid).pack(),
            )
        )
    kb.adjust(1)

    left_cb = UserCatalogCb(action="site_page", page=page - 1, supplier_id=supplier_id, product_id=0).pack() if page > 0 else None
    right_cb = UserCatalogCb(action="site_page", page=page + 1, supplier_id=supplier_id, product_id=0).pack() if page < pages - 1 else None
    nav = _nav_buttons(left_cb, right_cb)
    if nav:
        kb.row(*nav)

    kb.row(InlineKeyboardButton(text="🔗 Другая ссылка", callback_data=UserCatalogCb(action="site_add", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="🔎 Поиск по базе", callback_data=UserCatalogCb(action="search_start", page=0, supplier_id=supplier_id, product_id=0).pack()))
    kb.row(InlineKeyboardButton(text="🏢 Меню поставщика", callback_data=UserCatalogCb(action="supp_open", page=0, supplier_id=supplier_id, product_id=0).pack()))

    await _safe_render_call(call, text, kb)



@router.callback_query(IsApprovedUser(), UserCatalogCb.filter(F.action == "site_open"))
async def site_open(call: CallbackQuery, callback_data: UserCatalogCb, settings: Settings, state: FSMContext) -> None:
    await state.set_state(SiteAddForm.browsing)
    await _screen_product_card(
        call,
        settings,
        supplier_id=int(callback_data.supplier_id),
        product_id=int(callback_data.product_id),
        back_page=int(callback_data.page),
        back_mode="site",
    )


# -------------------- Screens: KP --------------------

async def _screen_kp(target: Message | CallbackQuery, settings: Settings, page: int = 0) -> None:
    tg_id = target.from_user.id
    total_items = await count_kp_items(settings.db_path, tg_id)

    if total_items == 0:
        text = "📄 <b>Моё КП</b>\n\nПока пусто. Добавьте товары из каталога."
        kb = InlineKeyboardBuilder()

        # ✅ кнопка должна быть даже при пустом КП
        kb.row(
            InlineKeyboardButton(
                text="🗑 Очистить КП",
                callback_data=UserKpCb(action="clear", page=0, item_id=0).pack(),
            )
        )

        kb.row(
            InlineKeyboardButton(
                text="📦 В каталог",
                callback_data=UserCatalogCb(action="supp_page", page=0, supplier_id=0, product_id=0).pack(),
            )
        )

        if isinstance(target, CallbackQuery):
            await _safe_render_call(target, text, kb)
        else:
            await target.answer(text, reply_markup=kb.as_markup())
        return

    # определим поставщиков в КП (для заголовка)
    all_items = await list_kp_items(settings.db_path, tg_id, limit=total_items, offset=0)

    supplier_ids = sorted({
        int(sid)
        for sid in (it.get("supplier_id") for it in all_items)
        if sid and str(sid).strip() and int(sid) > 0
    })

    if not supplier_ids:
        supp_line = "Поставщик: <b>—</b>"
    elif len(supplier_ids) == 1:
        supp_name = await _supplier_name(settings, supplier_ids[0])
        supp_line = f"Поставщик: <b>{html.escape(str(supp_name))}</b>"
    else:
        names: list[str] = []
        for sid in supplier_ids[:3]:
            names.append(str(await _supplier_name(settings, sid)))
        tail = "…" if len(supplier_ids) > 3 else ""
        supp_line = f"Поставщики: <b>{html.escape(', '.join(names) + tail)}</b>"

    page, pages = _clamp_page(page, total_items, KP_PAGE_SIZE)
    items = await list_kp_items(settings.db_path, tg_id, limit=KP_PAGE_SIZE, offset=page * KP_PAGE_SIZE)

    lines: list[str] = []
    base_index = page * KP_PAGE_SIZE
    for i, it in enumerate(items, start=1):
        idx = base_index + i
        title = _short(it.get("title") or it.get("description") or "—", 58)
        price = _money(it.get("final_price") or it.get("price"))
        lines.append(f"{idx}. {html.escape(title)}\n   Цена: <b>{html.escape(price)}</b>")

    text = f"📄 <b>Моё КП</b>\n{supp_line}\n\n" + "\n\n".join(lines)
    if pages > 1:
        text += f"\n\nСтр. <b>{page + 1}</b> / <b>{pages}</b>"

    kb = InlineKeyboardBuilder()
    for it in items:
        kb.row(
            InlineKeyboardButton(
                text=f"❌ Убрать: {_short(it.get('title') or 'позиция', 28)}",
                callback_data=UserKpCb(action="del", page=page, item_id=int(it["id"])).pack(),
            )
        )

    left_cb = UserKpCb(action="page", page=page - 1, item_id=0).pack() if page > 0 else None
    right_cb = UserKpCb(action="page", page=page + 1, item_id=0).pack() if page < pages - 1 else None
    nav = _nav_buttons(left_cb, right_cb)
    if nav:
        kb.row(*nav)

    kb.row(InlineKeyboardButton(text="🗑 Очистить всё", callback_data=UserKpCb(action="clear", page=0, item_id=0).pack()))
    kb.row(InlineKeyboardButton(text="📦 В каталог", callback_data=UserCatalogCb(action="supp_page", page=0, supplier_id=0, product_id=0).pack()))

    if isinstance(target, CallbackQuery):
        await _safe_render_call(target, text, kb)
    else:
        await target.answer(text, reply_markup=kb.as_markup())


@router.callback_query(IsApprovedUser(), UserKpCb.filter(F.action == "view"))
async def kp_view(call: CallbackQuery, callback_data: UserKpCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_kp(call, settings, page=int(callback_data.page))


@router.callback_query(IsApprovedUser(), UserKpCb.filter(F.action == "page"))
async def kp_page(call: CallbackQuery, callback_data: UserKpCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await _screen_kp(call, settings, page=int(callback_data.page))


@router.callback_query(IsApprovedUser(), UserKpCb.filter(F.action == "del"))
async def kp_del(call: CallbackQuery, callback_data: UserKpCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await remove_kp_item(settings.db_path, call.from_user.id, int(callback_data.item_id))
    await _screen_kp(call, settings, page=int(callback_data.page))


@router.callback_query(UserKpCb.filter(F.action == "clear"))
async def kp_clear(call: CallbackQuery, callback_data: UserKpCb, settings: Settings, state: FSMContext) -> None:
    await state.clear()
    await clear_kp(settings.db_path, call.from_user.id)

    try:
        await call.answer("КП очищено")
    except Exception:
        pass

    await _screen_kp(call, settings, page=0)


@router.callback_query(IsApprovedUser(), F.data == "noop")
async def noop(call: CallbackQuery) -> None:
    await call.answer("❌ Нет в наличии.", show_alert=True)
