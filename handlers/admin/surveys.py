from __future__ import annotations

import asyncio
import math
from datetime import datetime
from html import escape
from io import BytesIO
from typing import Any

from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile, CallbackQuery, InlineKeyboardButton, Message
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl import Workbook

from callbacks.surveys import SurveyAdminCb
from config import Settings
from db import is_admin, list_user_ids
from db.surveys import (
    count_surveys,
    create_survey,
    get_survey_export_rows,
    get_survey_options,
    get_survey_results,
    list_surveys,
    mark_survey_sent,
    save_survey_delivery,
)
from keyboards.admin import admin_back_cancel_kb, admin_main_kb
from keyboards.surveys import (
    admin_survey_result_kb,
    admin_surveys_list_kb,
    admin_surveys_menu_kb,
    survey_confirm_kb,
    survey_mode_kb,
    survey_preview_kb,
    survey_user_kb,
)

router = Router()

PAGE_SIZE = 7
BTN_SURVEYS = "🗳 Опросы"
SKIP_ATTACHMENT = "⏭ Без вложения"


class SurveyForm(StatesGroup):
    question = State()
    attachment = State()
    options = State()
    confirm = State()


def _safe(value: Any, default: str = "—") -> str:
    text = str(value or "").strip()
    return escape(text or default)


def _date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "—"
    return escape(text.replace("T", " ").replace("+00:00", " UTC"))


def _mode_title(mode: str) -> str:
    return "один вариант" if mode == "single" else "несколько вариантов"


def _parse_options(text: str) -> list[str]:
    items: list[str] = []
    seen: set[str] = set()

    for raw_line in (text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue

        for prefix in ("-", "•", "—", "–"):
            if line.startswith(prefix):
                line = line[len(prefix):].strip()

        if len(line) > 3 and line[0].isdigit():
            line = line.lstrip("0123456789").lstrip(".) ").strip()

        line = " ".join(line.split())
        if not line:
            continue

        key = line.casefold()
        if key in seen:
            continue

        seen.add(key)
        items.append(line[:120])

    return items


def _attachment_kb():
    from aiogram.types import KeyboardButton, ReplyKeyboardMarkup

    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=SKIP_ATTACHMENT)],
            [KeyboardButton(text="⬅️ Назад"), KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
        input_field_placeholder="Фото, файл или пропуск",
    )


async def _check_admin(message_or_call: Message | CallbackQuery, settings: Settings) -> bool:
    allowed = await is_admin(settings.db_path, message_or_call.from_user.id, settings.admin_ids)
    if allowed:
        return True

    if isinstance(message_or_call, CallbackQuery):
        await message_or_call.answer("Нет доступа", show_alert=True)
    else:
        await message_or_call.answer("Нет доступа")
    return False


async def _send_survey_content(
    bot,
    chat_id: int,
    question: str,
    mode: str,
    options: list[dict],
    survey_id: int,
    file_id: str | None = None,
    file_kind: str | None = None,
) -> None:
    if file_id and file_kind == "photo":
        await bot.send_photo(chat_id=chat_id, photo=file_id, caption="🖼 Материал к опросу")
    elif file_id and file_kind == "document":
        await bot.send_document(chat_id=chat_id, document=file_id, caption="📎 Материал к опросу")

    text = (
        "🗳 <b>Новый опрос</b>\n\n"
        f"<b>{escape(question)}</b>\n\n"
        f"Можно выбрать: <b>{_mode_title(mode)}</b>."
    )

    await bot.send_message(
        chat_id=chat_id,
        text=text,
        reply_markup=survey_user_kb(
            survey_id=survey_id,
            mode=mode,
            options=options,
        ),
    )


async def _send_preview(message: Message, data: dict[str, Any]) -> None:
    file_id = data.get("file_id")
    file_kind = data.get("file_kind")
    question = str(data.get("question") or "")
    mode = str(data.get("mode") or "single")
    options = list(data.get("options") or [])

    await message.answer("👁 <b>Предпросмотр опроса</b>:")

    if file_id and file_kind == "photo":
        await message.answer_photo(file_id, caption="🖼 Материал к опросу")
    elif file_id and file_kind == "document":
        await message.answer_document(file_id, caption="📎 Материал к опросу")

    await message.answer(
        "🗳 <b>Новый опрос</b>\n\n"
        f"<b>{escape(question)}</b>\n\n"
        f"Можно выбрать: <b>{_mode_title(mode)}</b>.",
        reply_markup=survey_preview_kb(mode=mode, options=options),
    )
    await message.answer("Отправить опрос всем пользователям?", reply_markup=survey_confirm_kb())


def _menu_text() -> str:
    return (
        "🗳 <b>Опросы</b>\n\n"
        "Здесь можно создать новый опрос для пользователей и посмотреть ответы по старым опросам.\n\n"
        "Опрос поддерживает один или несколько вариантов ответа, фото и файлы."
    )


async def _send_menu(message: Message) -> None:
    await message.answer(_menu_text(), reply_markup=admin_surveys_menu_kb())


async def _show_surveys_list_message(message: Message, settings: Settings, page: int) -> None:
    total = await count_surveys(settings.db_path)
    total_pages = max(1, math.ceil(total / PAGE_SIZE))
    page = max(1, min(page, total_pages))
    items = await list_surveys(settings.db_path, limit=PAGE_SIZE, offset=(page - 1) * PAGE_SIZE)

    text = (
        "📊 <b>Результаты опросов</b>\n\n"
        f"Всего опросов: <b>{total}</b>\n"
        f"Страница: <b>{page}/{total_pages}</b>\n\n"
    )
    if not items:
        text += "Пока нет отправленных опросов."
    else:
        text += "Выберите опрос для просмотра подробной статистики."

    await message.answer(text, reply_markup=admin_surveys_list_kb(items, page, total_pages))


async def _show_surveys_list_call(call: CallbackQuery, settings: Settings, page: int) -> None:
    total = await count_surveys(settings.db_path)
    total_pages = max(1, math.ceil(total / PAGE_SIZE))
    page = max(1, min(page, total_pages))
    items = await list_surveys(settings.db_path, limit=PAGE_SIZE, offset=(page - 1) * PAGE_SIZE)

    text = (
        "📊 <b>Результаты опросов</b>\n\n"
        f"Всего опросов: <b>{total}</b>\n"
        f"Страница: <b>{page}/{total_pages}</b>\n\n"
    )
    if not items:
        text += "Пока нет отправленных опросов."
    else:
        text += "Выберите опрос для просмотра подробной статистики."

    await call.message.edit_text(text, reply_markup=admin_surveys_list_kb(items, page, total_pages))


def _results_text(result: dict[str, Any]) -> str:
    voters = int(result.get("voters_count") or 0)
    delivered = int(result.get("delivered_count") or result.get("sent_count") or 0)
    failed = int(result.get("delivery_failed_count") or result.get("failed_count") or 0)
    choices = int(result.get("choices_count") or 0)

    text = (
        "📊 <b>Результаты опроса</b>\n\n"
        f"<b>{_safe(result.get('question'))}</b>\n\n"
        f"Формат: <b>{_mode_title(str(result.get('select_mode') or 'single'))}</b>\n"
        f"Отправлено: <b>{delivered}</b>\n"
        f"Ошибки отправки: <b>{failed}</b>\n"
        f"Ответили: <b>{voters}</b>\n"
        f"Всего выбранных вариантов: <b>{choices}</b>\n"
        f"Дата отправки: <b>{_date(result.get('sent_at'))}</b>\n\n"
        "<b>Варианты:</b>\n"
    )

    options = result.get("options") or []
    if not options:
        text += "—"
        return text

    base = voters if str(result.get("select_mode")) == "single" else max(choices, 1)
    for index, option in enumerate(options, start=1):
        votes = int(option.get("votes_count") or 0)
        percent = 0 if base <= 0 else round(votes * 100 / base, 1)
        bar_len = min(10, max(0, round(percent / 10)))
        bar = "█" * bar_len + "░" * (10 - bar_len)
        text += f"\n{index}. {_safe(option.get('text'))}\n{bar} <b>{votes}</b> · {percent}%\n"

    return text.strip()


def _fit_sheet_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 90)


def _build_excel_report(data: dict[str, Any]) -> BufferedInputFile:
    survey = data["survey"]
    responses = data["responses"]
    not_answered = data["not_answered"]

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Итоги"
    ws_summary.append(["Вопрос", survey.get("question")])
    ws_summary.append(["Формат", _mode_title(str(survey.get("select_mode") or "single"))])
    ws_summary.append(["Отправлено", int(survey.get("delivered_count") or 0)])
    ws_summary.append(["Ошибки отправки", int(survey.get("delivery_failed_count") or 0)])
    ws_summary.append(["Ответили", int(survey.get("voters_count") or 0)])
    ws_summary.append(["Дата отправки", survey.get("sent_at") or ""])
    ws_summary.append([])
    ws_summary.append(["Вариант", "Ответов"])
    for option in survey.get("options") or []:
        ws_summary.append([option.get("text") or "", int(option.get("votes_count") or 0)])
    _fit_sheet_columns(ws_summary)

    ws_answers = wb.create_sheet("Ответы")
    ws_answers.append(["Telegram ID", "ФИО", "Телефон", "Ответ", "Дата ответа"])
    for row in responses:
        ws_answers.append([
            row.get("tg_id"),
            row.get("full_name") or "",
            row.get("phone") or "",
            row.get("selected_options") or "",
            row.get("created_at") or "",
        ])
    _fit_sheet_columns(ws_answers)

    ws_no = wb.create_sheet("Без ответа")
    ws_no.append(["Telegram ID", "ФИО", "Телефон", "Статус доставки", "Ошибка", "Дата"])
    for row in not_answered:
        ws_no.append([
            row.get("tg_id"),
            row.get("full_name") or "",
            row.get("phone") or "",
            row.get("status") or "",
            row.get("error") or "",
            row.get("created_at") or "",
        ])
    _fit_sheet_columns(ws_no)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"survey_{survey.get('id')}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return BufferedInputFile(output.getvalue(), filename=filename)


@router.message(F.text.in_({BTN_SURVEYS, "Опросы", "🗳 Опросы"}))
async def surveys_menu_message(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    await state.clear()
    await _send_menu(message)


@router.callback_query(SurveyAdminCb.filter(F.action == "menu"))
async def surveys_menu_call(call: CallbackQuery, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(call, settings):
        return

    await state.clear()
    await call.message.edit_text(_menu_text(), reply_markup=admin_surveys_menu_kb())
    await call.answer()


@router.callback_query(SurveyAdminCb.filter(F.action == "back"))
async def surveys_back(call: CallbackQuery, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(call, settings):
        return

    await state.clear()
    await call.message.answer("🛠 <b>Админ-панель</b>", reply_markup=admin_main_kb())
    try:
        await call.message.delete()
    except Exception:
        pass
    await call.answer()


@router.callback_query(SurveyAdminCb.filter(F.action == "noop"))
async def surveys_noop(call: CallbackQuery) -> None:
    await call.answer("Предпросмотр")


@router.callback_query(SurveyAdminCb.filter(F.action == "cancel"))
async def surveys_cancel_call(call: CallbackQuery, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(call, settings):
        return

    await state.clear()
    await call.message.answer("❌ Создание опроса отменено.", reply_markup=admin_main_kb())
    try:
        await call.message.delete()
    except Exception:
        pass
    await call.answer()


@router.callback_query(SurveyAdminCb.filter(F.action == "create"))
async def surveys_create_start(call: CallbackQuery, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(call, settings):
        return

    await state.clear()
    await state.set_state(SurveyForm.question)
    await call.message.answer(
        "➕ <b>Новый опрос</b>\n\n"
        "Напишите вопрос для пользователей.\n\n"
        "Пример: <i>Какой формат работы удобнее?</i>",
        reply_markup=admin_back_cancel_kb(),
    )
    await call.answer()


@router.message(SurveyForm.question, F.text == "❌ Отмена")
@router.message(SurveyForm.attachment, F.text == "❌ Отмена")
@router.message(SurveyForm.options, F.text == "❌ Отмена")
@router.message(SurveyForm.confirm, F.text == "❌ Отмена")
async def surveys_cancel_message(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    await state.clear()
    await message.answer("❌ Создание опроса отменено.", reply_markup=admin_main_kb())


@router.message(SurveyForm.question, F.text == "⬅️ Назад")
@router.message(SurveyForm.attachment, F.text == "⬅️ Назад")
@router.message(SurveyForm.options, F.text == "⬅️ Назад")
@router.message(SurveyForm.confirm, F.text == "⬅️ Назад")
async def surveys_back_message(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    await state.clear()
    await message.answer("🛠 <b>Админ-панель</b>", reply_markup=admin_main_kb())


@router.message(SurveyForm.question, F.text)
async def surveys_question_take(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    question = " ".join((message.text or "").strip().split())
    if len(question) < 5:
        await message.answer("Вопрос слишком короткий. Напишите подробнее.", reply_markup=admin_back_cancel_kb())
        return

    if len(question) > 900:
        await message.answer("Вопрос слишком длинный. Уложитесь в 900 символов.", reply_markup=admin_back_cancel_kb())
        return

    await state.update_data(question=question)
    await message.answer(
        "Как пользователь сможет ответить?",
        reply_markup=survey_mode_kb(),
    )


@router.callback_query(SurveyAdminCb.filter(F.action == "mode"))
async def surveys_mode_take(
    call: CallbackQuery,
    callback_data: SurveyAdminCb,
    state: FSMContext,
    settings: Settings,
) -> None:
    if not await _check_admin(call, settings):
        return

    mode = callback_data.mode if callback_data.mode in {"single", "multiple"} else "single"
    await state.update_data(mode=mode)
    await state.set_state(SurveyForm.attachment)
    await call.message.answer(
        "📎 <b>Вложение к опросу</b>\n\n"
        "Отправьте фото или файл.\n"
        "Если вложение не нужно — нажмите «⏭ Без вложения».",
        reply_markup=_attachment_kb(),
    )
    await call.answer()


@router.message(SurveyForm.attachment, F.text == SKIP_ATTACHMENT)
async def surveys_skip_attachment(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    await state.update_data(file_id=None, file_kind=None)
    await state.set_state(SurveyForm.options)
    await message.answer(
        "🧩 <b>Варианты ответа</b>\n\n"
        "Напишите варианты одним сообщением, каждый вариант с новой строки.\n\n"
        "Пример:\n"
        "Да\n"
        "Нет\n"
        "Нужно обсудить",
        reply_markup=admin_back_cancel_kb(),
    )


@router.message(SurveyForm.attachment, F.photo)
async def surveys_take_photo(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    await state.update_data(file_id=message.photo[-1].file_id, file_kind="photo")
    await state.set_state(SurveyForm.options)
    await message.answer(
        "🖼 Фото добавлено.\n\n"
        "Теперь напишите варианты ответа одним сообщением, каждый вариант с новой строки.",
        reply_markup=admin_back_cancel_kb(),
    )


@router.message(SurveyForm.attachment, F.document)
async def surveys_take_document(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    await state.update_data(file_id=message.document.file_id, file_kind="document")
    await state.set_state(SurveyForm.options)
    await message.answer(
        "📎 Файл добавлен.\n\n"
        "Теперь напишите варианты ответа одним сообщением, каждый вариант с новой строки.",
        reply_markup=admin_back_cancel_kb(),
    )


@router.message(SurveyForm.attachment)
async def surveys_attachment_invalid(message: Message) -> None:
    await message.answer("Отправьте фото, файл или нажмите «⏭ Без вложения».", reply_markup=_attachment_kb())


@router.message(SurveyForm.options, F.text)
async def surveys_options_take(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(message, settings):
        return

    options = _parse_options(message.text or "")
    if len(options) < 2:
        await message.answer("Нужно минимум 2 варианта ответа. Напишите каждый вариант с новой строки.")
        return

    if len(options) > 12:
        await message.answer("Слишком много вариантов. Оставьте до 12 вариантов, чтобы кнопки выглядели аккуратно.")
        return

    await state.update_data(options=options)
    await state.set_state(SurveyForm.confirm)

    data = await state.get_data()
    await _send_preview(message, data)


@router.message(SurveyForm.options)
async def surveys_options_invalid(message: Message) -> None:
    await message.answer("Напишите варианты текстом: каждый вариант с новой строки.")


@router.callback_query(SurveyAdminCb.filter(F.action == "send"))
async def surveys_send_all(call: CallbackQuery, state: FSMContext, settings: Settings) -> None:
    if not await _check_admin(call, settings):
        return

    data = await state.get_data()
    question = str(data.get("question") or "").strip()
    mode = str(data.get("mode") or "single")
    file_id = data.get("file_id")
    file_kind = data.get("file_kind")
    options_text = list(data.get("options") or [])

    if not question or mode not in {"single", "multiple"} or len(options_text) < 2:
        await call.answer("Не хватает данных для отправки.", show_alert=True)
        return

    survey_id = await create_survey(
        settings.db_path,
        question=question,
        select_mode=mode,
        file_id=file_id,
        file_kind=file_kind,
        created_by=call.from_user.id,
        options=options_text,
    )
    options = await get_survey_options(settings.db_path, survey_id)
    user_ids = await list_user_ids(settings.db_path, exclude_statuses=("blocked",))

    status_msg = await call.message.answer(
        "⏳ <b>Отправляю опрос</b>\n\n"
        f"Получателей: <b>{len(user_ids)}</b>"
    )

    ok = 0
    fail = 0

    for tg_id in user_ids:
        try:
            await _send_survey_content(
                bot=call.bot,
                chat_id=tg_id,
                question=question,
                mode=mode,
                options=options,
                survey_id=survey_id,
                file_id=file_id,
                file_kind=file_kind,
            )
            ok += 1
            await save_survey_delivery(settings.db_path, survey_id, tg_id, "sent")
        except Exception as e:
            fail += 1
            await save_survey_delivery(settings.db_path, survey_id, tg_id, "failed", str(e)[:500])

        await asyncio.sleep(0.05)

    await mark_survey_sent(settings.db_path, survey_id, ok, fail)
    await state.clear()

    try:
        await status_msg.edit_text(
            "✅ <b>Опрос отправлен</b>\n\n"
            f"Получили: <b>{ok}</b>\n"
            f"Ошибки: <b>{fail}</b>\n\n"
            "Результаты появятся в разделе «Опросы»."
        )
    except Exception:
        pass

    await call.message.answer("🛠 <b>Админ-панель</b>", reply_markup=admin_main_kb())
    await call.answer()


@router.callback_query(SurveyAdminCb.filter(F.action == "list"))
async def surveys_list_call(
    call: CallbackQuery,
    callback_data: SurveyAdminCb,
    settings: Settings,
) -> None:
    if not await _check_admin(call, settings):
        return

    await _show_surveys_list_call(call, settings, callback_data.page)
    await call.answer()


@router.callback_query(SurveyAdminCb.filter(F.action == "open"))
async def surveys_open_result(
    call: CallbackQuery,
    callback_data: SurveyAdminCb,
    settings: Settings,
) -> None:
    if not await _check_admin(call, settings):
        return

    result = await get_survey_results(settings.db_path, callback_data.survey_id)
    if not result:
        await call.answer("Опрос не найден", show_alert=True)
        return

    await call.message.edit_text(
        _results_text(result),
        reply_markup=admin_survey_result_kb(callback_data.survey_id, callback_data.page),
    )
    await call.answer()


@router.callback_query(SurveyAdminCb.filter(F.action == "export"))
async def surveys_export_result(
    call: CallbackQuery,
    callback_data: SurveyAdminCb,
    settings: Settings,
) -> None:
    if not await _check_admin(call, settings):
        return

    data = await get_survey_export_rows(settings.db_path, callback_data.survey_id)
    if not data:
        await call.answer("Опрос не найден", show_alert=True)
        return

    report_file = _build_excel_report(data)
    await call.message.answer_document(
        document=report_file,
        caption="📄 <b>Excel-отчёт по опросу</b>",
    )
    await call.answer("Отчёт готов")
