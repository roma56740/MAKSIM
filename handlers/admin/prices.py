from __future__ import annotations

import json
import logging
import math
import os
import re
import tempfile
import hashlib
from typing import Any

from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, InlineKeyboardButton, Message
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl import load_workbook

from callbacks import PricesCb
from config import Settings
from db import (
    count_suppliers,
    get_supplier,
    is_admin,
    list_suppliers,
)
from db.catalog import (
    upsert_product_by_code,
    upsert_supplier_price,
    count_products,
    list_products,
)
from keyboards.admin import admin_back_cancel_kb, admin_main_kb

try:
    import xlrd  # для .xls
except Exception:
    xlrd = None  # если не установлено


router = Router()
PAGE_SIZE = 8
log = logging.getLogger(__name__)


class PriceUpload(StatesGroup):
    waiting_excel = State()


# -------------------- basic helpers --------------------

def _clean_text(v: Any) -> str | None:
    s = str(v).strip() if v is not None else ""
    s = re.sub(r"\s+", " ", s)
    return s or None


def _safe_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("₽", "").replace("руб.", "").replace("руб", "").replace("\xa0", " ")
    s = s.replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s or s in {"-", "."}:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _safe_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        try:
            return int(v)
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d\-]", "", s)
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def _safe_code(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if float(v).is_integer():
            return str(int(v))
        return str(v).strip()
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s or None


def _calc_final_price(price: float | None, discount_percent: float | None, final_price: float | None) -> float | None:
    if final_price is not None:
        return float(final_price)
    if price is None:
        return None
    if discount_percent is None:
        return float(price)
    try:
        fp = float(price) * (1.0 - float(discount_percent) / 100.0)
        return round(fp, 2)
    except Exception:
        return float(price)


def _parse_volume(v: Any) -> float | None:
    """
    Возвращает объём в литрах.
    Понимает:
      - 0.75, 0,75
      - 750 ml
      - 0.75 l / 0,75 л
      - 700 (как ml)
    """
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        x = float(v)
        if x <= 0:
            return None
        if x >= 20:
            return round(x / 1000.0, 4)
        return x

    s = str(v).strip().lower().replace(",", ".")
    if not s:
        return None

    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:l|л)\b", s)
    if m:
        try:
            x = float(m.group(1))
            return x if 0 < x <= 10 else None
        except Exception:
            pass

    m = re.search(r"(\d{2,4})\s*ml\b", s)
    if m:
        try:
            return round(int(m.group(1)) / 1000.0, 4)
        except Exception:
            pass

    if re.fullmatch(r"\d+(?:\.\d+)?", s):
        try:
            x = float(s)
            return x if 0 < x <= 10 else None
        except Exception:
            return None

    return None


def _money(v: Any) -> str:
    if v is None or v == "":
        return "—"
    try:
        return f"{float(v):,.2f}".replace(",", " ").replace(".", ",")
    except Exception:
        return str(v)


# -------------------- AUTO EXCEL PARSER --------------------

def _norm_header(s: Any) -> str:
    t = str(s).strip().lower() if s is not None else ""
    t = t.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    t = re.sub(r"\s+", " ", t)
    t = t.replace("ё", "е")
    return t


_FIELD_RULES: dict[str, dict[str, list[str] | int]] = {
    "code": {
        "pos": [r"\bкод\b", r"код товара", r"\bsku\b", r"\bid\b", r"item code", r"product id"],
        "neg": [r"штрих", r"\bean\b", r"barcode"],
        "w": 3,
    },
    "source_pk": {
        "pos": [r"артик", r"\bарт\.?\b", r"article", r"vendor code", r"арт\."],
        "neg": [],
        "w": 3,
    },
    "barcode": {
        "pos": [r"штрих", r"\bean\b", r"barcode"],
        "neg": [],
        "w": 3,
    },
    "title": {
        "pos": [r"наимен", r"номенклат", r"\bname\b", r"описан", r"\bproduct\b", r"название"],
        "neg": [r"\bкод\b", r"\bid\b", r"\bsku\b", r"штрих", r"\bean\b", r"barcode", r"цена", r"скидк", r"остат", r"\bтип\b"],
        "w": 3,
    },
    "strength": {
        "pos": [r"креп", r"\babv\b", r"градус", r"dosage", r"дозаж", r"\bтип\b", r"category", r"категор"],
        "neg": [r"цена", r"скидк", r"остат", r"штрих", r"url", r"ссылка", r"фото", r"image"],
        "w": 2,
    },
    "volume": {
        "pos": [r"объем", r"объ[eе]м", r"емк", r"\bvolume\b", r"\bml\b", r"\bl\b", r"литр"],
        "neg": [r"цена", r"скидк", r"остат", r"штрих", r"url", r"фото"],
        "w": 3,
    },
    "price": {
        "pos": [r"\bцена\b", r"\bprice\b", r"стоим"],
        "neg": [r"скидк", r"клиент", r"финал", r"со скид", r"discount"],
        "w": 3,
    },
    "discount_percent": {
        "pos": [r"процент", r"\b%\b", r"скид", r"discount"],
        "neg": [r"цена"],
        "w": 2,
    },
    "final_price": {
        "pos": [r"цена для клиента", r"цена со скид", r"финал", r"final", r"со скид", r"sale price"],
        "neg": [],
        "w": 3,
    },
    "stock_qty": {
        "pos": [r"остат", r"налич", r"свободн", r"\bqty\b", r"quantity", r"\bstock\b", r"balance"],
        "neg": [r"штрих", r"\bean\b", r"barcode"],
        "w": 3,
    },
    "url": {
        "pos": [r"\burl\b", r"ссылка", r"link"],
        "neg": [],
        "w": 3,
    },
    "image_url": {
        "pos": [r"фото", r"\bimage\b", r"picture", r"photo", r"картин", r"изображ"],
        "neg": [],
        "w": 3,
    },
}


def _col_field_score(h: str, field: str) -> int:
    rules = _FIELD_RULES[field]
    score = 0
    w = int(rules["w"])  # type: ignore[arg-type]
    for p in rules["pos"]:  # type: ignore[assignment]
        if re.search(p, h):
            score += w
    for n in rules["neg"]:  # type: ignore[assignment]
        if re.search(n, h):
            score -= w * 2
    return score


def _map_columns(headers: list[str]) -> tuple[dict[str, int], list[str]]:
    hn = [_norm_header(h) for h in headers]
    col: dict[str, int] = {}

    for field in _FIELD_RULES.keys():
        best_i: int | None = None
        best_s = -10**9

        for i, h in enumerate(hn):
            if not h:
                continue
            s = _col_field_score(h, field)
            if s > best_s:
                best_s = s
                best_i = i
            elif s == best_s and s > 0 and best_i is not None:
                if field == "title":
                    pref = ("наимен" in h) or (h == "номенклатура")
                    pref_best = ("наимен" in hn[best_i]) or (hn[best_i] == "номенклатура")
                    if pref and not pref_best:
                        best_i = i

        if best_i is not None and best_s > 0:
            col[field] = best_i

    if "title" not in col:
        for i, h in enumerate(hn):
            if "описан" in h or "description" in h:
                col["title"] = i
                break

    return col, hn


def _header_row_score(headers: list[str]) -> int:
    col, _ = _map_columns(headers)
    key = sum(1 for k in ("code", "title", "price", "final_price", "stock_qty") if k in col)
    bonus = 0
    if ("price" in col or "final_price" in col) and ("title" in col or "code" in col or "source_pk" in col):
        bonus = 2
    return key * 10 + bonus


def _build_code_fallback(
    *,
    supplier_id: int,
    title: str | None,
    source_pk: str | None,
    barcode: str | None,
    volume: float | None,
    strength: str | None
) -> str:
    base = "|".join([
        str(supplier_id),
        source_pk or "",
        barcode or "",
        title or "",
        strength or "",
        str(volume or ""),
    ])
    h = hashlib.sha1(base.encode("utf-8")).hexdigest()[:12]
    return f"AUTO-{h}"


def _parse_rows_auto_from_xlsx(path: str, supplier_id: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)

    best: dict[str, Any] | None = None

    for sname in wb.sheetnames:
        ws = wb[sname]
        best_row: dict[str, Any] | None = None

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx > 200:
                break
            vals = [str(v).strip() if v is not None else "" for v in row]
            non_empty = sum(1 for v in vals if v)
            if non_empty < 3:
                continue

            score = _header_row_score(vals)
            if score >= 18:
                if best_row is None or score > best_row["score"]:
                    col, _ = _map_columns(vals)
                    best_row = {
                        "sheet": sname,
                        "header_row": r_idx,
                        "headers": vals,
                        "columns": col,
                        "score": score,
                    }

        if best_row is None:
            continue

        quality = int(best_row["score"]) + len(best_row["columns"]) * 2
        if best is None or quality > int(best["quality"]):
            best = {**best_row, "quality": quality}

    if best is None:
        wb.close()
        raise ValueError("Не удалось определить строку заголовков. Проверьте, что в файле есть колонки (название/цена/код/остаток).")

    ws = wb[best["sheet"]]
    headers: list[str] = best["headers"]
    columns: dict[str, int] = best["columns"]
    headers_norm = [_norm_header(h) for h in headers]

    rows: list[dict[str, Any]] = []

    it = ws.iter_rows(values_only=True)
    for _ in range(int(best["header_row"])):
        next(it, None)

    empty_streak = 0

    def _get(row_: tuple[Any, ...], field: str) -> Any:
        idx = columns.get(field)
        if idx is None:
            return None
        return row_[idx] if idx < len(row_) else None

    for row in it:
        if row is None:
            continue

        if all(v is None or str(v).strip() == "" for v in row):
            empty_streak += 1
            if empty_streak >= 25 and rows:
                break
            continue
        empty_streak = 0

        non_empty_idx = [i for i, v in enumerate(row) if v is not None and str(v).strip() != ""]
        if len(non_empty_idx) == 1 and non_empty_idx[0] == 0 and isinstance(row[0], str):
            continue

        code = _safe_code(_get(row, "code"))
        source_pk = _clean_text(_get(row, "source_pk"))
        barcode = _safe_code(_get(row, "barcode"))
        title = _clean_text(_get(row, "title"))
        strength = _clean_text(_get(row, "strength"))
        volume = _parse_volume(_get(row, "volume"))

        price = _safe_float(_get(row, "price"))
        discount_percent = _safe_float(_get(row, "discount_percent"))
        final_price = _safe_float(_get(row, "final_price"))
        stock_qty = _safe_int(_get(row, "stock_qty"))
        url = _clean_text(_get(row, "url"))
        image_url = _clean_text(_get(row, "image_url"))

        if not image_url:
            for i, h in enumerate(headers_norm):
                if "фото" in h or "image" in h or "picture" in h or "photo" in h:
                    if i < len(row):
                        v = _clean_text(row[i])
                        if v:
                            image_url = v
                            break

        has_id = bool(title or code or source_pk)
        has_nums = any(x is not None for x in (price, final_price, stock_qty, volume))
        if not has_id or not has_nums:
            continue

        if discount_percent is None and price is not None and final_price is not None and price > 0:
            try:
                discount_percent = round((1.0 - float(final_price) / float(price)) * 100.0, 4)
            except Exception:
                pass

        if price is None and final_price is not None:
            price = float(final_price)

        final_price = _calc_final_price(price, discount_percent, final_price)

        if not code:
            code = barcode or source_pk or _build_code_fallback(
                supplier_id=supplier_id,
                title=title,
                source_pk=source_pk,
                barcode=barcode,
                volume=volume,
                strength=strength,
            )

        rows.append({
            "code": code,
            "source_pk": source_pk,
            "barcode": barcode,
            "title": title,
            "strength": strength,
            "volume": volume,
            "price": price,
            "discount_percent": discount_percent,
            "final_price": final_price,
            "stock_qty": stock_qty,
            "url": url,
            "image_url": image_url,
        })

    wb.close()

    meta = {
        "engine": "xlsx/openpyxl",
        "sheet": best["sheet"],
        "header_row": int(best["header_row"]),
        "columns": best["columns"],  # field -> index
    }
    return rows, meta


def _parse_rows_auto_from_xls(path: str, supplier_id: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if xlrd is None:
        raise RuntimeError("Для чтения .xls нужен пакет xlrd. Установите: pip install xlrd")

    book = xlrd.open_workbook(path)
    best: dict[str, Any] | None = None

    for si in range(book.nsheets):
        sh = book.sheet_by_index(si)
        best_row: dict[str, Any] | None = None

        scan = min(200, sh.nrows)
        for r in range(scan):
            vals_raw = sh.row_values(r)
            vals = [str(v).strip() if v is not None else "" for v in vals_raw]
            non_empty = sum(1 for v in vals if v)
            if non_empty < 3:
                continue

            score = _header_row_score(vals)
            if score >= 18:
                if best_row is None or score > best_row["score"]:
                    col, _ = _map_columns(vals)
                    best_row = {
                        "sheet": sh.name,
                        "sheet_index": si,
                        "header_row": r + 1,  # 1-based
                        "headers": vals,
                        "columns": col,
                        "score": score,
                    }

        if best_row is None:
            continue

        quality = int(best_row["score"]) + len(best_row["columns"]) * 2
        if best is None or quality > int(best["quality"]):
            best = {**best_row, "quality": quality}

    if best is None:
        raise ValueError("Не удалось определить строку заголовков в .xls. Проверьте, что есть колонки (название/цена/код/остаток).")

    sh = book.sheet_by_index(int(best["sheet_index"]))
    headers: list[str] = best["headers"]
    columns: dict[str, int] = best["columns"]
    headers_norm = [_norm_header(h) for h in headers]

    rows: list[dict[str, Any]] = []

    def _get(row_: list[Any], field: str) -> Any:
        idx = columns.get(field)
        if idx is None:
            return None
        return row_[idx] if idx < len(row_) else None

    empty_streak = 0
    start = int(best["header_row"])  # 1-based
    for r in range(start, sh.nrows):
        row = sh.row_values(r)

        if all(v is None or str(v).strip() == "" for v in row):
            empty_streak += 1
            if empty_streak >= 25 and rows:
                break
            continue
        empty_streak = 0

        non_empty_idx = [i for i, v in enumerate(row) if v is not None and str(v).strip() != ""]
        if len(non_empty_idx) == 1 and non_empty_idx[0] == 0 and isinstance(row[0], str):
            continue

        code = _safe_code(_get(row, "code"))
        source_pk = _clean_text(_get(row, "source_pk"))
        barcode = _safe_code(_get(row, "barcode"))
        title = _clean_text(_get(row, "title"))
        strength = _clean_text(_get(row, "strength"))
        volume = _parse_volume(_get(row, "volume"))

        price = _safe_float(_get(row, "price"))
        discount_percent = _safe_float(_get(row, "discount_percent"))
        final_price = _safe_float(_get(row, "final_price"))
        stock_qty = _safe_int(_get(row, "stock_qty"))
        url = _clean_text(_get(row, "url"))
        image_url = _clean_text(_get(row, "image_url"))

        if not image_url:
            for i, h in enumerate(headers_norm):
                if "фото" in h or "image" in h or "picture" in h or "photo" in h:
                    if i < len(row):
                        v = _clean_text(row[i])
                        if v:
                            image_url = v
                            break

        has_id = bool(title or code or source_pk)
        has_nums = any(x is not None for x in (price, final_price, stock_qty, volume))
        if not has_id or not has_nums:
            continue

        if discount_percent is None and price is not None and final_price is not None and price > 0:
            try:
                discount_percent = round((1.0 - float(final_price) / float(price)) * 100.0, 4)
            except Exception:
                pass

        if price is None and final_price is not None:
            price = float(final_price)

        final_price = _calc_final_price(price, discount_percent, final_price)

        if not code:
            code = barcode or source_pk or _build_code_fallback(
                supplier_id=supplier_id,
                title=title,
                source_pk=source_pk,
                barcode=barcode,
                volume=volume,
                strength=strength,
            )

        rows.append({
            "code": code,
            "source_pk": source_pk,
            "barcode": barcode,
            "title": title,
            "strength": strength,
            "volume": volume,
            "price": price,
            "discount_percent": discount_percent,
            "final_price": final_price,
            "stock_qty": stock_qty,
            "url": url,
            "image_url": image_url,
        })

    meta = {
        "engine": "xls/xlrd",
        "sheet": best["sheet"],
        "header_row": int(best["header_row"]),
        "columns": best["columns"],
    }
    return rows, meta


def parse_products_auto_excel(path: str, supplier_id: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    low = path.lower()
    if low.endswith(".xlsx"):
        return _parse_rows_auto_from_xlsx(path, supplier_id)
    if low.endswith(".xls"):
        return _parse_rows_auto_from_xls(path, supplier_id)
    raise ValueError("Поддерживаются только .xlsx или .xls")


# -------------------- UI helpers --------------------

def _suppliers_list_text(items: list[dict], page: int, total_pages: int, total: int) -> str:
    if not items:
        body = "Пока нет поставщиков. Сначала добавьте поставщика в разделе «🏢 Поставщики»."
    else:
        lines = []
        n0 = page * PAGE_SIZE + 1
        for i, s in enumerate(items, start=n0):
            lines.append(f"{i}. <b>{s['name']}</b>  (ID <code>{s['id']}</code>)")
        body = "\n".join(lines)

    return (
        "📦 <b>Excel-прайсы поставщиков</b>\n"
        f"Всего поставщиков: <b>{total}</b>\n"
        f"Страница: <b>{page + 1}/{total_pages}</b>\n\n"
        f"{body}\n\n"
        "Выберите поставщика ниже."
    )


def _suppliers_list_kb(page: int, total_pages: int, items: list[dict]) -> InlineKeyboardBuilder:
    kb = InlineKeyboardBuilder()
    for s in items:
        kb.add(
            InlineKeyboardButton(
                text=f"📌 {s['name']}",
                callback_data=PricesCb(action="view", page=page, supplier_id=int(s["id"]), mode="").pack(),
            )
        )
    kb.adjust(1)

    prev_page = page - 1 if page > 0 else 0
    next_page = page + 1 if page + 1 < total_pages else page

    kb.row(
        InlineKeyboardButton(text="⬅️", callback_data=PricesCb(action="page", page=prev_page, supplier_id=0, mode="").pack()),
        InlineKeyboardButton(text="➡️", callback_data=PricesCb(action="page", page=next_page, supplier_id=0, mode="").pack()),
    )
    kb.row(
        InlineKeyboardButton(text="⬅️ В меню", callback_data=PricesCb(action="back", page=page, supplier_id=0, mode="").pack()),
    )
    return kb


def _products_list_text(supplier_name: str, items: list[dict], page: int, total_pages: int, total: int) -> str:
    if not items:
        body = "Пока нет товаров у этого поставщика."
    else:
        lines = []
        for p in items:
            title = p.get("title") or p.get("description") or "—"
            code = p.get("code") or "—"
            price = _money(p.get("final_price") or p.get("price"))
            stock = p.get("stock_qty")
            stock_s = str(stock) if stock is not None else "—"
            lines.append(f"• <b>{title}</b>\n  код: <code>{code}</code> | цена: <b>{price}</b> | остаток: <b>{stock_s}</b>")
        body = "\n\n".join(lines)

    return (
        "👀 <b>Товары поставщика</b>\n"
        f"🏢 <b>{supplier_name}</b>\n"
        f"Всего: <b>{total}</b>\n"
        f"Страница: <b>{page + 1}/{total_pages}</b>\n\n"
        f"{body}"
    )


def _products_list_kb(supplier_id: int, page: int, total_pages: int) -> InlineKeyboardBuilder:
    kb = InlineKeyboardBuilder()

    prev_page = page - 1 if page > 0 else 0
    next_page = page + 1 if page + 1 < total_pages else page

    kb.row(
        InlineKeyboardButton(text="⬅️", callback_data=PricesCb(action="list", page=prev_page, supplier_id=supplier_id, mode="").pack()),
        InlineKeyboardButton(text="➡️", callback_data=PricesCb(action="list", page=next_page, supplier_id=supplier_id, mode="").pack()),
    )
    kb.row(
        InlineKeyboardButton(text="⬅️ Назад", callback_data=PricesCb(action="view", page=0, supplier_id=supplier_id, mode="").pack()),
    )
    return kb


async def _render_suppliers(call_or_msg: Message | CallbackQuery, settings: Settings, page: int, edit: bool) -> None:
    total = await count_suppliers(settings.db_path)
    total_pages = max(1, math.ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))

    items = await list_suppliers(settings.db_path, limit=PAGE_SIZE, offset=page * PAGE_SIZE)
    text = _suppliers_list_text(items, page, total_pages, total)
    kb = _suppliers_list_kb(page, total_pages, items).as_markup()

    if isinstance(call_or_msg, CallbackQuery):
        if edit:
            await call_or_msg.message.edit_text(text, reply_markup=kb)
        else:
            await call_or_msg.message.answer(text, reply_markup=kb)
    else:
        await call_or_msg.answer(text, reply_markup=kb)


# -------------------- UI entry --------------------

@router.message(F.text == "📦 Excel-прайсы")
async def prices_open(message: Message, settings: Settings, state: FSMContext) -> None:
    if not await is_admin(settings.db_path, message.from_user.id, settings.admin_ids):
        return
    await state.clear()
    await _render_suppliers(message, settings, page=0, edit=False)


@router.callback_query(PricesCb.filter(F.action == "page"))
async def prices_page(call: CallbackQuery, callback_data: PricesCb, settings: Settings) -> None:
    if not await is_admin(settings.db_path, call.from_user.id, settings.admin_ids):
        await call.answer("Нет доступа", show_alert=True)
        return
    await _render_suppliers(call, settings, page=callback_data.page, edit=True)
    await call.answer()


@router.callback_query(PricesCb.filter(F.action == "back"))
async def prices_back(call: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    try:
        await call.message.delete()
    except Exception:
        pass
    await call.message.answer("🛠 <b>Админ-панель</b>", reply_markup=admin_main_kb())
    await call.answer()


@router.callback_query(PricesCb.filter(F.action == "view"))
async def prices_supplier_view(call: CallbackQuery, callback_data: PricesCb, settings: Settings) -> None:
    if not await is_admin(settings.db_path, call.from_user.id, settings.admin_ids):
        await call.answer("Нет доступа", show_alert=True)
        return

    s = await get_supplier(settings.db_path, callback_data.supplier_id)
    if not s:
        await call.answer("Поставщик не найден", show_alert=True)
        return

    kb = InlineKeyboardBuilder()
    kb.row(
        InlineKeyboardButton(
            text="👀 Посмотреть товары",
            callback_data=PricesCb(action="list", page=0, supplier_id=int(s["id"]), mode="").pack(),
        )
    )
    kb.row(
        InlineKeyboardButton(
            text="📥 Загрузить Excel",
            callback_data=PricesCb(action="upload", page=callback_data.page, supplier_id=int(s["id"]), mode="").pack(),
        )
    )
    kb.row(
        InlineKeyboardButton(
            text="⬅️ Назад",
            callback_data=PricesCb(action="page", page=callback_data.page, supplier_id=0, mode="").pack(),
        )
    )

    await call.message.answer(
        "📦 <b>Excel-прайс поставщика</b>\n\n"
        f"🏢 <b>{s['name']}</b>\n"
        f"🆔 <code>{s['id']}</code>\n\n"
        "Выберите действие:",
        reply_markup=kb.as_markup(),
    )
    await call.answer()


@router.callback_query(PricesCb.filter(F.action == "list"))
async def prices_supplier_list(call: CallbackQuery, callback_data: PricesCb, settings: Settings) -> None:
    if not await is_admin(settings.db_path, call.from_user.id, settings.admin_ids):
        await call.answer("Нет доступа", show_alert=True)
        return

    s = await get_supplier(settings.db_path, callback_data.supplier_id)
    if not s:
        await call.answer("Поставщик не найден", show_alert=True)
        return

    total = await count_products(settings.db_path, int(s["id"]))
    total_pages = max(1, math.ceil(total / PAGE_SIZE))
    page = max(0, min(int(callback_data.page), total_pages - 1))

    items = await list_products(settings.db_path, int(s["id"]), limit=PAGE_SIZE, offset=page * PAGE_SIZE)
    text = _products_list_text(str(s["name"]), items, page, total_pages, total)
    kb = _products_list_kb(int(s["id"]), page, total_pages).as_markup()

    await call.message.answer(text, reply_markup=kb)
    await call.answer()


@router.callback_query(PricesCb.filter(F.action == "upload"))
async def prices_upload_choose_mode(call: CallbackQuery, callback_data: PricesCb, settings: Settings) -> None:
    if not await is_admin(settings.db_path, call.from_user.id, settings.admin_ids):
        await call.answer("Нет доступа", show_alert=True)
        return

    kb = InlineKeyboardBuilder()
    kb.row(
        InlineKeyboardButton(
            text="➕ Добавить недостающие",
            callback_data=PricesCb(action="mode", page=callback_data.page, supplier_id=callback_data.supplier_id, mode="add_missing").pack(),
        )
    )
    kb.row(
        InlineKeyboardButton(
            text="♻️ Изменить данные и добавить",
            callback_data=PricesCb(action="mode", page=callback_data.page, supplier_id=callback_data.supplier_id, mode="upsert").pack(),
        )
    )
    kb.row(
        InlineKeyboardButton(
            text="⬅️ Назад",
            callback_data=PricesCb(action="view", page=callback_data.page, supplier_id=callback_data.supplier_id, mode="").pack(),
        )
    )

    await call.message.answer(
        "📥 <b>Загрузка Excel</b>\n\n"
        "Выберите режим:\n"
        "➕ Добавить недостающие — существующие товары не трогаем\n"
        "♻️ Изменить данные и добавить — обновляем существующие + добавляем новые\n\n"
        "Далее отправьте <b>.xlsx</b> или <b>.xls</b> файлом.",
        reply_markup=kb.as_markup(),
    )
    await call.answer()


@router.callback_query(PricesCb.filter(F.action == "mode"))
async def prices_upload_wait_file(call: CallbackQuery, callback_data: PricesCb, state: FSMContext, settings: Settings) -> None:
    if not await is_admin(settings.db_path, call.from_user.id, settings.admin_ids):
        await call.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await state.update_data(
        supplier_id=callback_data.supplier_id,
        mode=callback_data.mode,
        return_page=callback_data.page,
    )
    await state.set_state(PriceUpload.waiting_excel)

    await call.message.answer(
        "📄 Пришлите Excel файлом (<b>.xlsx</b> или <b>.xls</b>).\n"
        "Отмена — кнопкой ниже.",
        reply_markup=admin_back_cancel_kb(),
    )
    await call.answer()


@router.message(PriceUpload.waiting_excel, F.text.in_(["❌ Отмена", "⬅️ Назад"]))
async def prices_upload_cancel(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("Ок.", reply_markup=admin_main_kb())


# -------------------- IMPORT: EXCEL -> DB --------------------

@router.message(PriceUpload.waiting_excel, F.document)
async def prices_upload_file(message: Message, state: FSMContext, settings: Settings) -> None:
    if not await is_admin(settings.db_path, message.from_user.id, settings.admin_ids):
        return

    data = await state.get_data()
    supplier_id = int(data["supplier_id"])
    mode = str(data["mode"])
    overwrite = mode == "upsert"

    doc = message.document
    name = doc.file_name or "price.xlsx"
    low = name.lower()

    if not (low.endswith(".xlsx") or low.endswith(".xls")):
        await message.answer("❌ Нужен файл <b>.xlsx</b> или <b>.xls</b>.")
        return

    await message.answer("⏳ Загружаю и автоматически определяю формат...")

    bot = message.bot
    ext = ".xlsx" if low.endswith(".xlsx") else ".xls"

    # ВАЖНО: сохраняем под фиксированным именем (без кириллицы/пробелов) — меньше проблем на Windows
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, f"upload{ext}")
        await bot.download(doc, destination=path)

        try:
            rows, meta = parse_products_auto_excel(path, supplier_id=supplier_id)
        except Exception as e:
            logging.exception("Excel parse error")
            await message.answer(f"❌ Ошибка чтения Excel: <code>{type(e).__name__}: {e}</code>")
            return

    await upsert_supplier_price(
        settings.db_path,
        supplier_id=supplier_id,
        tg_file_id=doc.file_id,
        file_name=name,
        uploaded_by=message.from_user.id,
    )

    present_fields = set((meta.get("columns") or {}).keys())
    cols = ", ".join(sorted(present_fields))
    await message.answer(
        "✅ <b>Формат определён автоматически</b>\n"
        f"Лист: <b>{meta.get('sheet')}</b>\n"
        f"Строка заголовков: <b>{meta.get('header_row')}</b>\n"
        f"Колонки: <code>{cols or '—'}</code>\n\n"
        f"📦 Начинаю обработку: <b>{len(rows)}</b> строк(и). Режим: <b>{mode}</b>"
    )

    created = updated = skipped = errors = 0
    first_error: str | None = None

    # ВАЖНО:
    # - если в файле НЕТ колонки (например, "цена"), мы НЕ трогаем цену в БД
    # - если колонка есть, но значение пустое — в режиме upsert это очистит поле (это логично)
    for i, r in enumerate(rows, start=1):
        try:
            code = _clean_text(r.get("code"))
            if not code:
                raise ValueError("Пустой код товара (не удалось построить)")

            title = _clean_text(r.get("title"))
            strength = _clean_text(r.get("strength"))
            volume = _parse_volume(r.get("volume"))
            source_pk = _clean_text(r.get("source_pk"))
            barcode = _clean_text(r.get("barcode"))
            url = _clean_text(r.get("url"))
            image_url = _clean_text(r.get("image_url"))

            price = _safe_float(r.get("price"))
            discount = _safe_float(r.get("discount_percent"))
            final_price = _safe_float(r.get("final_price"))
            stock_qty = _safe_int(r.get("stock_qty"))
            if stock_qty is not None and stock_qty < 0:
                stock_qty = None

            # совместимость со старым выводом
            description = title
            product_type = strength

            payload: dict[str, Any] = {
                "code": code,
            }

            # добавляем только те поля, которые реально есть в файле (по meta.columns)
            if "source_pk" in present_fields:
                payload["source_pk"] = source_pk

            if "title" in present_fields:
                payload["title"] = title
                payload["description"] = description

            if "strength" in present_fields:
                payload["strength"] = strength
                payload["product_type"] = product_type

            if "volume" in present_fields:
                payload["volume"] = volume

            if "stock_qty" in present_fields:
                payload["stock_qty"] = stock_qty

            if "url" in present_fields:
                payload["url"] = url

            if "image_url" in present_fields:
                payload["image_url"] = image_url
                payload["image_path"] = None

            # цены: если в файле есть хотя бы одна из ценовых колонок — обновляем ценовой блок
            price_fields_present = bool({"price", "final_price", "discount_percent"} & present_fields)
            if price_fields_present:
                if discount is None and price is not None and final_price is not None and price > 0:
                    try:
                        discount = round((1.0 - float(final_price) / float(price)) * 100.0, 4)
                    except Exception:
                        pass

                if price is None and final_price is not None:
                    price = float(final_price)

                fp = _calc_final_price(price, discount, final_price)

                if "price" in present_fields or "final_price" in present_fields or "discount_percent" in present_fields:
                    payload["price"] = price
                    payload["discount_percent"] = discount
                    payload["final_price"] = fp

            payload["extra_json"] = json.dumps(
                {
                    "barcode": barcode,
                    "detected_sheet": meta.get("sheet"),
                    "header_row": meta.get("header_row"),
                    "present_fields": sorted(present_fields),
                },
                ensure_ascii=False
            )

            status, _pid = await upsert_product_by_code(
                settings.db_path,
                supplier_id=supplier_id,
                code=code,
                data=payload,
                overwrite=overwrite,
            )

            if status == "created":
                created += 1
            elif status == "updated":
                updated += 1
            else:
                skipped += 1

        except Exception as e:
            errors += 1
            logging.exception("Import error row=%s code=%s", i, r.get("code", "?"))
            if first_error is None:
                first_error = f"Строка {i}, код {r.get('code', '?')}: {type(e).__name__}: {e}"

    await state.clear()

    text = (
        "✅ <b>Готово</b>\n\n"
        f"➕ Добавлено: <b>{created}</b>\n"
        f"♻️ Обновлено: <b>{updated}</b>\n"
        f"⏭ Пропущено: <b>{skipped}</b>\n"
        f"⚠️ Ошибок: <b>{errors}</b>\n"
    )
    if first_error:
        text += f"\n<b>Первая ошибка:</b>\n<code>{first_error}</code>"

    await message.answer(text, reply_markup=admin_main_kb())
